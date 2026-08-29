from pathlib import Path

from textwrap import dedent

import base64
import json
import re
import time

import requests
import pandas as pd
import streamlit as st
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

from datetime import date, datetime
from io import BytesIO
from urllib.parse import urlencode
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image


# =========================================================
# CONFIGURACIÓN GENERAL
# =========================================================

# En local usa tu backend en 127.0.0.1:8000. En Streamlit Cloud, define
# la variable "URL_BACKEND" en Settings > Secrets con la URL de Render,
# por ejemplo: URL_BACKEND = "https://gastro-biopsias-backend.onrender.com"
try:
    URL_BACKEND = st.secrets["URL_BACKEND"]
except Exception:
    URL_BACKEND = "http://127.0.0.1:8000"

FORM_ID_FIJO = "1i0mNhszKOCrZN0AqvoCb5QuFnyDL7gNwio6algry_pU"
FORM_URL_EDITAR_FIJO = (
    f"https://docs.google.com/forms/d/{FORM_ID_FIJO}/edit"
)
FORM_URL_RESPONDER_FIJO = (
    f"https://docs.google.com/forms/d/{FORM_ID_FIJO}/viewform"
)

CARPETA_FRONTEND = Path(__file__).resolve().parent
CARPETA_PROYECTO = CARPETA_FRONTEND.parent

RUTA_LOGO = CARPETA_FRONTEND / "logo.png"
RUTA_LOGO_IZQUIERDA = CARPETA_FRONTEND / "logo_gobrlima.png"
RUTA_LOGO_DERECHA = CARPETA_FRONTEND / "logo_hmatucana.png"
RUTA_FONDO_HOSPITAL = CARPETA_FRONTEND / "fhospm.png"
RUTA_CREDENCIALES = CARPETA_PROYECTO / "credentials.json"

SCOPES_GOOGLE = [
    "https://www.googleapis.com/auth/forms.body",
    "https://www.googleapis.com/auth/forms.responses.readonly",
    "https://www.googleapis.com/auth/drive.file",
]

TITULO_GOOGLE_FORM = "Registro de Biopsias de Gastroenterología"

# =========================================================
# APPS SCRIPT - CREACIÓN DEL GOOGLE FORM
# =========================================================
# Pegue aquí la URL del despliegue Web App de Code.gs.
# Debe terminar normalmente en /exec

URL_APPS_SCRIPT = "https://script.google.com/macros/s/AKfycbxyw3nCq7ArPti5pwP6B1eSUkNAsqOJjJjFoBRT38jlZuC-xXUtc4QREozUsyb5Iafe/exec"

# Debe ser la misma clave configurada en Code.gs.
CLAVE_APPS_SCRIPT = "gastro-biopsias-2026"

# Si logo.png no existe, intenta buscar estomago.png
if not RUTA_LOGO.exists():
    RUTA_LOGO = CARPETA_FRONTEND / "estomago.png"

ICONO_PAGINA = str(RUTA_LOGO) if RUTA_LOGO.exists() else "🩺"


def obtener_logo_transparente(ruta_logo):
    """
    Quita únicamente el fondo parecido al color de la esquina del PNG.
    Conserva el dibujo del logo y devuelve una imagen PNG en memoria.
    """
    if not ruta_logo.exists():
        return None

    imagen = Image.open(ruta_logo).convert("RGBA")
    fondo = imagen.getpixel((0, 0))

    # Si ya tiene transparencia real, se usa tal cual.
    if fondo[3] < 20:
        salida = BytesIO()
        imagen.save(salida, format="PNG")
        salida.seek(0)
        return salida.getvalue()

    fr, fg, fb, _ = fondo
    pixeles = []

    for r, g, b, a in imagen.getdata():
        distancia = (
            (r - fr) ** 2 +
            (g - fg) ** 2 +
            (b - fb) ** 2
        ) ** 0.5

        # Solo elimina colores muy parecidos al fondo original.
        if distancia < 42:
            pixeles.append((r, g, b, 0))
        else:
            pixeles.append((r, g, b, a))

    imagen.putdata(pixeles)

    salida = BytesIO()
    imagen.save(salida, format="PNG")
    salida.seek(0)

    return salida.getvalue()



def obtener_imagen_base64(ruta_imagen):
    """
    Convierte una imagen local a base64 para usarla como fondo CSS.
    """
    if not ruta_imagen.exists():
        return ""

    return base64.b64encode(
        ruta_imagen.read_bytes()
    ).decode("utf-8")


FONDO_HOSPITAL_BASE64 = obtener_imagen_base64(
    RUTA_FONDO_HOSPITAL
)


st.set_page_config(
    page_title="GastroEnterología",
    page_icon=ICONO_PAGINA,
    layout="wide",
    initial_sidebar_state="expanded",
)


# =========================================================
# CAMPOS DEL FORMULARIO
# =========================================================

# La columna "Firma" NO se crea en Google Forms.
# Se agrega únicamente al reporte/Excel y siempre queda en blanco
# para la firma física posterior.
#
# N.º Biopsia ahora contiene dos cuadrículas:
#   1) Procedimiento y cantidad: cada procedimiento aparece en una fila.
#   2) Biopsia y cantidad: cada zona aparece en una fila.
# La cantidad se marca en la misma fila, evitando preguntas separadas debajo.
PROCEDIMIENTOS_BIOPSIA = [
    "Endoscopia",
    "Colonoscopia",
    "Sigmoidoscopia",
    "Proctoscopia",
    "Otros",
]

BIOPSIAS_POR_PROCEDIMIENTO = {
    "Endoscopia": [
        "Antro",
        "Cuerpo",
        "Ángulo",
        "Otros",
    ],
    "Colonoscopia": [
        "Recto",
        "Colon ascendente",
        "Colon transverso",
        "Pólipo",
        "Otros",
    ],
    "Sigmoidoscopia": [
        "Sigmoides",
        "Otros",
    ],
    "Proctoscopia": [
        "Recto",
        "Otros",
    ],

}

# Lista general usada para crear el combo de Google Forms.
# Se mantiene una sola pregunta Biopsia para que el reporte sea simple.
CATEGORIAS_BIOPSIA = [
    "Antro",
    "Cuerpo",
    "Ángulo",
    "Recto",
    "Colon ascendente",
    "Colon transverso",
    "Pólipo",
    "Otros",
]

# Columnas del cuaderno físico (Sedación/Anestesia hasta Mucosectomía).
# Igual que Biopsia, es una cuadrícula: cada técnica en una fila, cantidad
# al costado.
PROCEDIMIENTOS_ADICIONALES = [
    "Sedación",
    "Anestesia",
    "APL",
    "ELVE",
    "Colocación enema",
    "Inyectoterapia",
    "Clip",
    "Polipectomía alta",
    "Polipectomía baja",
    "Mucosectomía",
]

PREFIJO_CANTIDAD_PROCEDIMIENTO = "Cantidad procedimiento - "
PREFIJO_CANTIDAD_BIOPSIA = "Cantidad biopsia - "
PREFIJO_CANTIDAD_ADICIONAL = "Cantidad adicional - "

# Columnas que se mostrarán en las cuadrículas de Google Forms.
# Deje la fila en blanco cuando ese procedimiento/biopsia no corresponda.
CANTIDADES_GRID = [str(numero) for numero in range(1, 11)]

CAMPOS_FORMULARIO = [
    {
        "titulo": "Fecha",
        "tipo": "fecha",
        "obligatorio": True,
    },
    {
        "titulo": "N.º Biopsia",
        "tipo": "biopsia",
        "procedimientos": PROCEDIMIENTOS_BIOPSIA,
        "biopsias_por_procedimiento": BIOPSIAS_POR_PROCEDIMIENTO,
        "categorias": CATEGORIAS_BIOPSIA,
        "adicionales": PROCEDIMIENTOS_ADICIONALES,
        # Procedimiento ahora es selección única + cantidad aparte
        # (ya no es cuadrícula). Biopsia y cantidad y Adicionales siguen
        # siendo cuadrícula.
        "procedimiento_cuadricula": False,
        "cantidad_por_fila": True,
        "cantidades": CANTIDADES_GRID,
        "obligatorio": False,
    },
    {
        "titulo": "Médico",
        "tipo": "parrafo",
        "obligatorio": True,
    },
    {
        "titulo": "Enfermera",
        "tipo": "parrafo",
        "obligatorio": True,
    },
    {
        "titulo": "Técnica",
        "tipo": "parrafo",
        "obligatorio": True,
    },
    {
        "titulo": "Observaciones",
        "tipo": "parrafo",
        "obligatorio": False,
    },
]


# =========================================================
# GOOGLE FORMS
# =========================================================

def obtener_credenciales_google():
    """
    Solicita autorización de Google usando credentials.json.

    No se guarda autorización en disco. Cada vez que una función
    necesite acceder a Google Forms, se abre la autorización en el navegador.
    """
    if not RUTA_CREDENCIALES.exists():
        raise FileNotFoundError(
            "No se encontró credentials.json en "
            f"{RUTA_CREDENCIALES}"
        )

    flujo = InstalledAppFlow.from_client_secrets_file(
        str(RUTA_CREDENCIALES),
        SCOPES_GOOGLE,
    )

    credenciales = flujo.run_local_server(
        port=0,
        open_browser=True,
        login_hint="servgastroenterologia446@gmail.com",
    )

    return credenciales


def crear_solicitud_campo(campo, indice):
    """
    Convierte un campo del sistema en un elemento de Google Forms.
    """
    titulo = campo["titulo"]
    tipo = campo["tipo"]
    obligatorio = bool(campo.get("obligatorio", False))

    item = {
        "title": titulo,
    }

    descripcion = str(campo.get("descripcion", "")).strip()
    if descripcion:
        item["description"] = descripcion

    # Encabezado visual: no genera una respuesta, solo agrupa
    # los campos Biopsia y Cantidad dentro del Google Form.
    if tipo == "encabezado":
        item["textItem"] = {}
        return {
            "createItem": {
                "item": item,
                "location": {
                    "index": indice,
                },
            }
        }

    pregunta = {
        "required": obligatorio,
    }

    if tipo == "fecha":
        pregunta["dateQuestion"] = {
            "includeYear": True,
            "includeTime": False,
        }

    elif tipo == "check":
        # Una sola opción permite usar el campo como casilla de verificación.
        # Marcado = "Sí"; sin marcar = respuesta vacía.
        pregunta["choiceQuestion"] = {
            "type": "CHECKBOX",
            "options": [
                {"value": "Sí"},
            ],
            "shuffle": False,
        }

    elif tipo == "lista":
        opciones = campo.get("opciones", [])

        if not opciones:
            raise ValueError(
                f"El campo '{titulo}' no tiene opciones."
            )

        pregunta["choiceQuestion"] = {
            "type": "DROP_DOWN",
            "options": [
                {"value": str(opcion)}
                for opcion in opciones
            ],
            "shuffle": False,
        }

    elif tipo == "radio_otro":
        # Radio button con la opción nativa "Otros" de Google Forms.
        # Al marcar Otros, Google habilita el ingreso de texto en la
        # misma pregunta, sin crear secciones ni partir el formulario.
        opciones = [
            str(opcion)
            for opcion in campo.get("opciones", [])
            if str(opcion).strip().casefold() != "otros"
        ]

        if not opciones:
            raise ValueError(
                f"El campo '{titulo}' no tiene opciones."
            )

        pregunta["choiceQuestion"] = {
            "type": "RADIO",
            "options": [
                {"value": opcion}
                for opcion in opciones
            ] + [
                {
                    # En una opción nativa "Otros", Google Forms
                    # no acepta enviar "value" junto con isOther=True.
                    # Google muestra automáticamente "Otros" y habilita
                    # el cuadro de texto al seleccionarlo.
                    "isOther": True,
                }
            ],
            "shuffle": False,
        }

    elif tipo == "texto":
        pregunta["textQuestion"] = {
            "paragraph": False,
        }

    elif tipo == "numero":
        item["description"] = "Ingrese únicamente números enteros."
        pregunta["textQuestion"] = {
            "paragraph": False,
        }

    elif tipo == "parrafo":
        pregunta["textQuestion"] = {
            "paragraph": True,
        }

    else:
        raise ValueError(
            f"Tipo de campo no reconocido: {tipo}"
        )

    item["questionItem"] = {
        "question": pregunta,
    }

    return {
        "createItem": {
            "item": item,
            "location": {
                "index": indice,
            },
        }
    }


def obtener_entry_id_google_form(
    enlace_respuestas,
    titulo_pregunta="Fecha",
    intentos=6,
):
    """
    Obtiene el ID numérico entry.* usado por los vínculos
    prellenados de Google Forms.

    Este ID no es igual al questionId que devuelve Forms API.
    """
    encabezados = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/126.0.0.0 Safari/537.36"
        )
    }

    ultimo_error = None

    for intento in range(intentos):
        try:
            respuesta = requests.get(
                enlace_respuestas,
                headers=encabezados,
                timeout=20,
            )
            respuesta.raise_for_status()

            coincidencia = re.search(
                r"var FB_PUBLIC_LOAD_DATA_ = (\[.+?\]);\s*<",
                respuesta.text,
                re.DOTALL,
            )

            if coincidencia is None:
                coincidencia = re.search(
                    r"var FB_PUBLIC_LOAD_DATA_ = (\[.+?\]);",
                    respuesta.text,
                    re.DOTALL,
                )

            if coincidencia is None:
                raise ValueError(
                    "Google todavía no publicó la estructura del formulario."
                )

            datos_publicos = json.loads(coincidencia.group(1))
            preguntas = datos_publicos[1][1] or []

            for pregunta in preguntas:
                if not pregunta or len(pregunta) < 5:
                    continue

                titulo = pregunta[1]
                detalles = pregunta[4] or []

                if titulo == titulo_pregunta and detalles:
                    entry_id = detalles[0][0]

                    if entry_id is not None:
                        return str(entry_id)

            raise ValueError(
                f"No se encontró el entry ID de '{titulo_pregunta}'."
            )

        except (
            requests.exceptions.RequestException,
            ValueError,
            IndexError,
            TypeError,
            json.JSONDecodeError,
        ) as error:
            ultimo_error = error

            if intento < intentos - 1:
                time.sleep(1.5)

    raise ValueError(
        f"No se pudo obtener el ID numérico del campo "
        f"{titulo_pregunta}. Detalle: {ultimo_error}"
    )

def crear_google_form(campos, progreso=None):
    """
    Crea el Google Form mediante Apps Script.

    Se usa Apps Script para crear las cuadrículas de Procedimiento y
    Biopsia, dejando cada opción en una fila y la cantidad en columnas.
    """
    if (
        not URL_APPS_SCRIPT
        or URL_APPS_SCRIPT == "PEGAR_AQUI_URL_WEB_APP"
        or not URL_APPS_SCRIPT.startswith("https://")
    ):
        raise ValueError(
            "Falta configurar URL_APPS_SCRIPT con la URL /exec "
            "del Web App de Google Apps Script."
        )

    print(f"[APPS SCRIPT] Creando formulario mediante: {URL_APPS_SCRIPT}")

    fecha_hoy_texto = datetime.now(
        ZoneInfo("America/Lima")
    ).strftime("%d/%m/%Y")

    titulo_formulario = (
        f"{TITULO_GOOGLE_FORM} - {fecha_hoy_texto}"
    )

    respuesta = requests.post(
        URL_APPS_SCRIPT,
        json={
            "clave": CLAVE_APPS_SCRIPT,
            "titulo": titulo_formulario,
            "campos": campos,
            # El Apps Script reutiliza el mismo formulario guardado si no
            # se fuerza uno nuevo. Como cada dia necesita un Google Form
            # con ID propio (asi lo espera SQLite), se pide siempre uno
            # nuevo aqui.
            "forzar_nuevo": True,
        },
        timeout=60,
    )

    respuesta.raise_for_status()

    if progreso is not None:
        progreso.progress(
            55,
            text="Google creó el formulario..."
        )

    try:
        datos = respuesta.json()
    except ValueError as error:
        raise ValueError(
            "Apps Script no devolvió una respuesta JSON válida."
        ) from error

    if not datos.get("ok"):
        raise ValueError(
            datos.get(
                "error",
                "Apps Script no pudo crear el Google Form.",
            )
        )

    form_id = datos["form_id"]
    enlace_respuestas = datos["enlace_respuestas"]
    enlace_edicion = datos["enlace_edicion"]

    if progreso is not None:
        progreso.progress(
            75,
            text="Configurando formulario..."
        )

    # Obtiene el entry.* del campo Fecha para conservar
    # el prellenado automático que ya usa la aplicación.
    entry_id_fecha = obtener_entry_id_google_form(
        enlace_respuestas,
        titulo_pregunta="Fecha",
    )

    if progreso is not None:
        progreso.progress(
            90,
            text="Finalizando formulario..."
        )

    return {
        "form_id": form_id,
        "titulo": titulo_formulario,
        "enlace_respuestas": enlace_respuestas,
        "enlace_edicion": enlace_edicion,
        "entry_id_fecha": entry_id_fecha,
    }


def convertir_fecha_creacion_formulario(valor):
    """
    Convierte fecha_creacion del backend a date.
    Soporta fecha simple, timestamp SQLite e ISO.
    """
    texto = str(valor or "").strip()

    if not texto:
        return None

    try:
        fecha_iso = datetime.fromisoformat(
            texto.replace("Z", "+00:00")
        )

        if fecha_iso.tzinfo is not None:
            fecha_iso = fecha_iso.astimezone(
                ZoneInfo("America/Lima")
            )

        return fecha_iso.date()
    except ValueError:
        pass

    for formato in (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y",
    ):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue

    return None


def obtener_formulario_hoy():
    """
    Intenta FastAPI. Si el backend responde pero no hay
    formulario creado hoy, devuelve None (para habilitar el
    boton de creacion). Solo usa el formulario fijo de Google
    cuando el backend esta realmente inalcanzable.
    """
    try:
        respuesta = requests.get(
            f"{URL_BACKEND}/formularios",
            timeout=5,
        )
        respuesta.raise_for_status()

        formularios = respuesta.json().get(
            "formularios",
            [],
        )

        hoy = datetime.now(
            ZoneInfo("America/Lima")
        ).date()

        for formulario in formularios:
            fecha_formulario = convertir_fecha_creacion_formulario(
                formulario.get("fecha_creacion")
            )

            if fecha_formulario == hoy:
                return formulario

        # El backend respondio correctamente pero no hay
        # formulario creado hoy.
        return None
    except Exception:
        # No se pudo contactar al backend: se usa el formulario fijo.
        return {
            "form_id_google": FORM_ID_FIJO,
            "titulo": TITULO_GOOGLE_FORM,
            "url_responder": FORM_URL_RESPONDER_FIJO,
            "url_editar": FORM_URL_EDITAR_FIJO,
            "fecha_creacion": datetime.now(
                ZoneInfo("America/Lima")
            ).strftime("%Y-%m-%d"),
        }


def cargar_formulario_hoy_en_sesion(formulario):
    """
    Recupera las URLs del formulario diario y las deja listas
    para los pasos 2 y 3, incluso después de reiniciar Streamlit.
    """
    if not formulario:
        return

    url_responder = str(
        formulario.get("url_responder", "") or ""
    ).strip()
    url_editar = str(
        formulario.get("url_editar", "") or ""
    ).strip()

    if url_responder:
        st.session_state["form_url_base"] = (
            url_responder.split("?")[0]
        )

        try:
            st.session_state["form_entry_id_fecha"] = (
                obtener_entry_id_google_form(
                    url_responder,
                    titulo_pregunta="Fecha",
                )
            )
        except Exception:
            # El formulario sigue siendo utilizable aunque, por algún
            # motivo puntual, no se pueda recuperar el entry de Fecha.
            st.session_state["form_entry_id_fecha"] = ""

    st.session_state["form_url_editar"] = url_editar
    st.session_state["form_fecha_activa"] = datetime.now(
        ZoneInfo("America/Lima")
    ).strftime("%Y-%m-%d")


# =========================================================
# REPORTES EN EXCEL
# =========================================================

MESES_ES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}

COLUMNAS_REPORTE = [
    "Fecha",
    "N.º Biopsia",
    "Médico",
    "Enfermera",
    "Técnica",
    "Observaciones",
    "Firma",
]


def convertir_fecha_formulario(valor):
    """
    Convierte el campo Fecha del formulario en date.
    """
    if isinstance(valor, date):
        return valor

    texto = str(valor or "").strip()

    for formato in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue

    return None


def convertir_fecha_envio_peru(valor):
    """
    Convierte la fecha de envío de Google Forms a hora de Perú.
    """
    texto = str(valor or "").strip()

    if not texto:
        return None

    try:
        fecha_hora = datetime.fromisoformat(
            texto.replace("Z", "+00:00")
        )

        if fecha_hora.tzinfo is None:
            fecha_hora = fecha_hora.replace(
                tzinfo=ZoneInfo("UTC")
            )

        return fecha_hora.astimezone(
            ZoneInfo("America/Lima")
        )

    except (TypeError, ValueError):
        return None


def obtener_todas_respuestas_google(servicio_forms, form_id):
    """
    Obtiene todas las respuestas, incluyendo formularios con paginación.
    """
    respuestas = []
    token = None

    while True:
        solicitud = servicio_forms.forms().responses().list(
            formId=form_id,
            pageToken=token,
        )

        resultado = solicitud.execute()
        respuestas.extend(resultado.get("responses", []))

        token = resultado.get("nextPageToken")
        if not token:
            break

    return respuestas


def obtener_registros_formulario(form_id):
    """
    Consulta las preguntas y respuestas del Google Form seleccionado.
    """
    credenciales = obtener_credenciales_google()

    servicio_forms = build(
        "forms",
        "v1",
        credentials=credenciales,
        cache_discovery=False,
    )

    datos_formulario = servicio_forms.forms().get(
        formId=form_id,
    ).execute()

    preguntas = {}
    orden_preguntas = []

    for item in datos_formulario.get("items", []):
        # Pregunta normal.
        pregunta = item.get(
            "questionItem",
            {},
        ).get(
            "question",
            {},
        )

        question_id = pregunta.get("questionId")

        if question_id:
            preguntas[question_id] = item.get(
                "title",
                question_id,
            )
            orden_preguntas.append(question_id)

        # Cuadrículas de Procedimiento/Biopsia. Google Forms expone
        # cada fila como una pregunta independiente (rowQuestion).
        grupo = item.get("questionGroupItem", {})
        titulo_grupo = str(item.get("title", "") or "").strip()

        if grupo:
            for pregunta_fila in grupo.get("questions", []):
                question_id_fila = pregunta_fila.get("questionId")
                titulo_fila = str(
                    pregunta_fila.get("rowQuestion", {}).get("title", "")
                    or ""
                ).strip()

                if not question_id_fila or not titulo_fila:
                    continue

                if titulo_grupo == "Procedimiento y cantidad":
                    titulo_respuesta = (
                        f"{PREFIJO_CANTIDAD_PROCEDIMIENTO}{titulo_fila}"
                    )
                elif titulo_grupo == "Biopsia y cantidad":
                    titulo_respuesta = (
                        f"{PREFIJO_CANTIDAD_BIOPSIA}{titulo_fila}"
                    )
                elif titulo_grupo == "Procedimientos adicionales y cantidad":
                    titulo_respuesta = (
                        f"{PREFIJO_CANTIDAD_ADICIONAL}{titulo_fila}"
                    )
                else:
                    titulo_respuesta = (
                        f"{titulo_grupo} - {titulo_fila}"
                        if titulo_grupo
                        else titulo_fila
                    )

                preguntas[question_id_fila] = titulo_respuesta
                orden_preguntas.append(question_id_fila)

    respuestas = obtener_todas_respuestas_google(
        servicio_forms,
        form_id,
    )

    registros = []

    for respuesta in respuestas:
        fecha_envio_texto = respuesta.get(
            "lastSubmittedTime",
            respuesta.get("createTime", ""),
        )

        fecha_envio_peru = convertir_fecha_envio_peru(
            fecha_envio_texto
        )

        registro = {
            "Fecha de envío": fecha_envio_texto,
        }

        respuestas_usuario = respuesta.get(
            "answers",
            {},
        )

        for question_id in orden_preguntas:
            titulo = preguntas[question_id]

            respuesta_pregunta = respuestas_usuario.get(
                question_id,
                {},
            )

            textos = respuesta_pregunta.get(
                "textAnswers",
                {},
            ).get(
                "answers",
                [],
            )

            valores = [
                texto.get("value", "")
                for texto in textos
                if str(texto.get("value", "")).strip()
            ]
            valor = ", ".join(valores)

            registro[titulo] = valor
            # Conserva la lista original para preguntas checkbox.
            # Así no se pierde qué opciones fueron marcadas.
            registro[f"_lista_{titulo}"] = valores

        # Reconstruye las selecciones de Procedimiento y Biopsia a partir
        # de las filas de la cuadrícula que sí tienen una cantidad marcada.
        procedimientos_seleccionados = [
            opcion
            for opcion in PROCEDIMIENTOS_BIOPSIA
            if str(
                registro.get(
                    f"{PREFIJO_CANTIDAD_PROCEDIMIENTO}{opcion}",
                    "",
                )
                or ""
            ).strip()
        ]
        biopsias_seleccionadas = [
            opcion
            for opcion in CATEGORIAS_BIOPSIA
            if str(
                registro.get(
                    f"{PREFIJO_CANTIDAD_BIOPSIA}{opcion}",
                    "",
                )
                or ""
            ).strip()
        ]
        adicionales_seleccionados = [
            opcion
            for opcion in PROCEDIMIENTOS_ADICIONALES
            if str(
                registro.get(
                    f"{PREFIJO_CANTIDAD_ADICIONAL}{opcion}",
                    "",
                )
                or ""
            ).strip()
        ]

        if procedimientos_seleccionados:
            registro["Procedimiento"] = ", ".join(
                procedimientos_seleccionados
            )
            registro["_lista_Procedimiento"] = procedimientos_seleccionados

        if biopsias_seleccionadas:
            registro["Biopsia"] = ", ".join(biopsias_seleccionadas)
            registro["_lista_Biopsia"] = biopsias_seleccionadas

        if adicionales_seleccionados:
            registro["ProcedimientosAdicionales"] = ", ".join(
                adicionales_seleccionados
            )
            registro["_lista_ProcedimientosAdicionales"] = (
                adicionales_seleccionados
            )

        registro["_fecha"] = convertir_fecha_formulario(
            registro.get("Fecha")
        )
        registro["_fecha_envio"] = fecha_envio_peru

        registros.append(registro)

    registros.sort(
        key=lambda registro: (
            registro.get("_fecha") or date.min,
            registro.get("_fecha_envio")
            or datetime.min.replace(
                tzinfo=ZoneInfo("America/Lima")
            ),
        )
    )

    return registros


def normalizar_check(valor):
    """
    Convierte la respuesta de una casilla en una marca visual.
    """
    texto = str(valor or "").strip().lower()

    if texto in {
        "sí",
        "si",
        "true",
        "1",
        "x",
        "✓",
        "check",
    }:
        return "✓"

    return ""


def obtener_selecciones_registro(registro, titulo):
    """
    Devuelve las opciones marcadas de una pregunta checkbox.
    En formularios nuevos se conserva la lista original. En formularios
    antiguos se intenta recuperar desde el texto separado por comas.
    """
    lista = registro.get(f"_lista_{titulo}")

    if isinstance(lista, list):
        return [str(valor).strip() for valor in lista if str(valor).strip()]

    texto = str(registro.get(titulo, "") or "").strip()
    if not texto:
        return []

    return [parte.strip() for parte in texto.split(",") if parte.strip()]


def tiene_cantidades_por_opcion_procedimiento(registro):
    """Indica si la respuesta usa la cuadrícula de Procedimiento
    (formularios antiguos). Los formularios nuevos usan selección única
    + cantidad, así que esto da False para ellos."""
    return any(
        str(clave).startswith(PREFIJO_CANTIDAD_PROCEDIMIENTO)
        for clave in registro.keys()
    )


def tiene_cantidades_por_opcion_biopsia(registro):
    """Indica si la respuesta usa la cuadrícula de Biopsia y cantidad."""
    return any(
        str(clave).startswith(PREFIJO_CANTIDAD_BIOPSIA)
        for clave in registro.keys()
    )


def tiene_cantidades_por_opcion_adicional(registro):
    """Indica si la respuesta usa la cuadrícula de Procedimientos
    adicionales y cantidad (APL, ELVE, enema, etc.)."""
    return any(
        str(clave).startswith(PREFIJO_CANTIDAD_ADICIONAL)
        for clave in registro.keys()
    )


def tiene_cantidades_por_opcion(registro):
    """Indica si la respuesta pertenece al nuevo formato de cantidades,
    ya sea en Procedimiento, en Biopsia, en Adicionales, o en varios."""
    return (
        tiene_cantidades_por_opcion_procedimiento(registro)
        or tiene_cantidades_por_opcion_biopsia(registro)
        or tiene_cantidades_por_opcion_adicional(registro)
    )


def obtener_items_con_cantidad(
    registro,
    titulo_seleccion,
    opciones_catalogo,
    prefijo_cantidad,
    campo_otro_compatibilidad=None,
):
    """
    Devuelve una lista de (nombre, cantidad) para procedimientos o biopsias.

    - Las opciones conocidas usan su propia pregunta de cantidad.
    - La opción nativa "Otros" de Google Forms devuelve normalmente el texto
      escrito por el usuario; ese texto usa la cantidad correspondiente a
      "Otros".
    - También recupera cantidades llenadas aunque el checkbox se haya omitido,
      para no perder datos en el reporte.
    """
    catalogo_fijo = [
        opcion for opcion in opciones_catalogo
        if str(opcion).casefold() != "otros"
    ]
    selecciones = obtener_selecciones_registro(registro, titulo_seleccion)
    vistos = set()
    items = []

    def agregar(nombre, cantidad):
        nombre = str(nombre or "").strip()
        cantidad = str(cantidad or "").strip()
        if not nombre:
            return
        clave = nombre.casefold()
        if clave in vistos:
            return
        vistos.add(clave)
        items.append((nombre, cantidad))

    for seleccion in selecciones:
        if seleccion in catalogo_fijo:
            agregar(
                seleccion,
                registro.get(f"{prefijo_cantidad}{seleccion}", ""),
            )
            continue

        # Puede venir literalmente "Otros" en versiones anteriores o puede
        # venir directamente el texto digitado en la opción nativa Otros.
        nombre_otro = seleccion
        if seleccion.casefold() == "otros" and campo_otro_compatibilidad:
            nombre_compatibilidad = str(
                registro.get(campo_otro_compatibilidad, "") or ""
            ).strip()
            if nombre_compatibilidad:
                nombre_otro = nombre_compatibilidad

        agregar(
            nombre_otro,
            registro.get(f"{prefijo_cantidad}Otros", ""),
        )

    # Si alguien escribió una cantidad pero olvidó marcar la casilla,
    # igualmente se conserva en los reportes.
    for opcion in catalogo_fijo:
        cantidad = str(
            registro.get(f"{prefijo_cantidad}{opcion}", "") or ""
        ).strip()
        if cantidad:
            agregar(opcion, cantidad)

    cantidad_otros = str(
        registro.get(f"{prefijo_cantidad}Otros", "") or ""
    ).strip()
    if cantidad_otros:
        nombre_otro = "Otros"
        if campo_otro_compatibilidad:
            nombre_compatibilidad = str(
                registro.get(campo_otro_compatibilidad, "") or ""
            ).strip()
            if nombre_compatibilidad:
                nombre_otro = nombre_compatibilidad
        agregar(nombre_otro, cantidad_otros)

    return items


def obtener_procedimientos_con_cantidad(registro):
    return obtener_items_con_cantidad(
        registro=registro,
        titulo_seleccion="Procedimiento",
        opciones_catalogo=PROCEDIMIENTOS_BIOPSIA,
        prefijo_cantidad=PREFIJO_CANTIDAD_PROCEDIMIENTO,
        campo_otro_compatibilidad="Otro procedimiento",
    )


def obtener_biopsias_con_cantidad(registro):
    return obtener_items_con_cantidad(
        registro=registro,
        titulo_seleccion="Biopsia",
        opciones_catalogo=CATEGORIAS_BIOPSIA,
        prefijo_cantidad=PREFIJO_CANTIDAD_BIOPSIA,
        campo_otro_compatibilidad="Otra biopsia",
    )


def obtener_adicionales_con_cantidad(registro):
    return obtener_items_con_cantidad(
        registro=registro,
        titulo_seleccion="ProcedimientosAdicionales",
        opciones_catalogo=PROCEDIMIENTOS_ADICIONALES,
        prefijo_cantidad=PREFIJO_CANTIDAD_ADICIONAL,
        campo_otro_compatibilidad=None,
    )


def resolver_procedimiento_biopsia(registro):
    """
    Devuelve el nombre del procedimiento seleccionado.

    Funciona tanto para el formulario nuevo (selección única "Procedimiento")
    como para formularios antiguos (checkbox de varias opciones). En ambos
    casos, si la opción elegida es "Otros", se sustituye por el texto
    escrito en "Otro procedimiento".
    """
    selecciones = obtener_selecciones_registro(registro, "Procedimiento")
    if selecciones:
        resueltas = []
        for seleccion in selecciones:
            if seleccion.casefold() == "otros":
                procedimiento_otro = str(
                    registro.get("Otro procedimiento", "") or ""
                ).strip()
                resueltas.append(procedimiento_otro or seleccion)
            else:
                resueltas.append(seleccion)
        return ", ".join(resueltas)

    procedimiento = str(registro.get("Procedimiento", "") or "").strip()
    if procedimiento.casefold() == "otros":
        procedimiento_otro = str(
            registro.get("Otro procedimiento", "") or ""
        ).strip()
        if procedimiento_otro:
            return procedimiento_otro
    return procedimiento


def resolver_nombre_biopsia(registro):
    """Compatibilidad con formularios antiguos de una sola cantidad."""
    selecciones = obtener_selecciones_registro(registro, "Biopsia")
    if selecciones:
        return ", ".join(selecciones)

    nombre = str(registro.get("Biopsia", "") or "").strip()
    if nombre.casefold() == "otros":
        nombre_otro = str(registro.get("Otra biopsia", "") or "").strip()
        if nombre_otro:
            return nombre_otro
    return nombre


def inferir_procedimiento_por_biopsia(nombre):
    """Compatibilidad con formularios antiguos."""
    nombre_limpio = str(nombre or "").strip()
    for procedimiento, categorias in BIOPSIAS_POR_PROCEDIMIENTO.items():
        categorias_sin_otros = [
            categoria for categoria in categorias if categoria != "Otros"
        ]
        if nombre_limpio in categorias_sin_otros:
            return procedimiento
    return ""


def formatear_items_cantidad(items):
    partes = []
    for nombre, cantidad in items:
        if cantidad:
            partes.append(f"{nombre} ({cantidad})")
        else:
            partes.append(nombre)
    return "; ".join(partes)


def formatear_biopsias(registro):
    """
    Construye la celda N.º Biopsia del Excel.

    Formato nuevo:
        Procedimientos: Endoscopia (2); Colonoscopia (1)
        Biopsias: Antro (3); Recto (1)

    Mantiene compatibilidad con formularios anteriores.
    """
    usa_grid_procedimiento = tiene_cantidades_por_opcion_procedimiento(
        registro
    )
    usa_grid_biopsia = tiene_cantidades_por_opcion_biopsia(registro)
    usa_grid_adicional = tiene_cantidades_por_opcion_adicional(registro)

    if usa_grid_procedimiento or usa_grid_biopsia or usa_grid_adicional:
        lineas = []

        if usa_grid_procedimiento:
            procedimientos = obtener_procedimientos_con_cantidad(registro)
            if procedimientos:
                lineas.append(
                    "Procedimientos: "
                    + formatear_items_cantidad(procedimientos)
                )
        else:
            # Formulario nuevo: Procedimiento es selección única + Cantidad.
            procedimiento = resolver_procedimiento_biopsia(registro)
            cantidad_procedimiento = str(
                registro.get("Cantidad", "") or ""
            ).strip()
            if procedimiento:
                if cantidad_procedimiento:
                    lineas.append(
                        f"Procedimiento: {procedimiento} "
                        f"({cantidad_procedimiento})"
                    )
                else:
                    lineas.append(f"Procedimiento: {procedimiento}")

        if usa_grid_biopsia:
            biopsias = obtener_biopsias_con_cantidad(registro)
            if biopsias:
                lineas.append(
                    "Biopsias: " + formatear_items_cantidad(biopsias)
                )

        if usa_grid_adicional:
            adicionales = obtener_adicionales_con_cantidad(registro)
            if adicionales:
                lineas.append(
                    "Adicionales: " + formatear_items_cantidad(adicionales)
                )

        if lineas:
            return "\n".join(lineas)

    # ---------------- FORMATO ANTERIOR ----------------
    procedimiento = resolver_procedimiento_biopsia(registro)
    nombre = resolver_nombre_biopsia(registro)
    cantidad = str(registro.get("Cantidad", "") or "").strip()

    if not procedimiento and nombre:
        procedimiento = inferir_procedimiento_por_biopsia(nombre)

    partes = []
    if procedimiento:
        partes.append(procedimiento)
    if nombre:
        partes.append(f"- {nombre}" if partes else nombre)

    texto = " ".join(partes).strip()
    if texto and cantidad:
        return f"{texto} ({cantidad})"
    if texto:
        return texto

    # Compatibilidad con la versión anterior de 3 pares.
    lineas = []
    for numero_biopsia in range(1, 4):
        nombre_anterior = str(
            registro.get(f"Biopsia {numero_biopsia} - Nombre", "") or ""
        ).strip()
        cantidad_anterior = str(
            registro.get(f"Biopsia {numero_biopsia} - Cantidad", "") or ""
        ).strip()

        if nombre_anterior:
            procedimiento_anterior = inferir_procedimiento_por_biopsia(
                nombre_anterior
            )
            prefijo = f"{procedimiento_anterior} - " if procedimiento_anterior else ""
            if cantidad_anterior:
                lineas.append(
                    f"{prefijo}{nombre_anterior} ({cantidad_anterior})"
                )
            else:
                lineas.append(f"{prefijo}{nombre_anterior}")

    # Compatibilidad con la estructura más antigua, una cantidad por zona.
    if not lineas:
        for categoria in CATEGORIAS_BIOPSIA:
            if categoria == "Otros":
                continue
            cantidad_anterior = str(
                registro.get(f"Biopsia - {categoria}", "") or ""
            ).strip()
            if cantidad_anterior:
                procedimiento_anterior = inferir_procedimiento_por_biopsia(
                    categoria
                )
                prefijo = f"{procedimiento_anterior} - " if procedimiento_anterior else ""
                lineas.append(
                    f"{prefijo}{categoria} ({cantidad_anterior})"
                )

    return "\n".join(lineas)

def convertir_cantidad_biopsia(valor):
    """
    Convierte una cantidad escrita como texto a entero para indicadores.
    Si no es un número entero válido, devuelve 0 sin alterar el reporte.
    """
    texto = str(valor or "").strip()

    if not texto:
        return 0

    try:
        return int(float(texto.replace(",", ".")))
    except (TypeError, ValueError):
        return 0


def construir_filas_reporte(registros):
    """
    Convierte cada respuesta en una fila del registro de biopsias.
    La última columna Firma se deja siempre en blanco.
    """
    filas = []

    for registro in registros:
        fecha = registro.get("_fecha")

        fila = {
            "Fecha": (
                fecha.strftime("%d/%m/%Y")
                if fecha
                else str(registro.get("Fecha", ""))
            ),
            "N.º Biopsia": formatear_biopsias(registro),
            "Médico": registro.get(
                "Médico",
                registro.get("Médico / Enfermera", ""),
            ),
            "Enfermera": registro.get(
                "Enfermera",
                "",
            ),
            "Técnica": registro.get(
                "Técnica",
                registro.get("Persona que recepciona", ""),
            ),
            "Observaciones": registro.get("Observaciones", ""),
            # Firma no se llena digitalmente: queda siempre en blanco
            # para la firma física posterior.
            "Firma": "",
        }

        filas.append(fila)

    return filas


def crear_excel_registro(filas, nombre_hoja):
    """
    Genera el Excel con el mismo orden de columnas del registro físico.
    Firma siempre queda en blanco.
    """
    libro = Workbook()
    hoja = libro.active
    hoja.title = nombre_hoja[:31]

    hoja.append(COLUMNAS_REPORTE)

    relleno_encabezado = PatternFill(
        "solid",
        fgColor="35566B",
    )
    fuente_encabezado = Font(
        color="FFFFFF",
        bold=True,
    )
    borde_fino = Border(
        left=Side(style="thin", color="7B8790"),
        right=Side(style="thin", color="7B8790"),
        top=Side(style="thin", color="7B8790"),
        bottom=Side(style="thin", color="7B8790"),
    )

    for celda in hoja[1]:
        celda.fill = relleno_encabezado
        celda.font = fuente_encabezado
        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        celda.border = borde_fino

    for fila in filas:
        hoja.append([
            fila.get(encabezado, "")
            for encabezado in COLUMNAS_REPORTE
        ])

        numero_fila = hoja.max_row

        for indice, celda in enumerate(
            hoja[numero_fila],
            start=1,
        ):
            celda.border = borde_fino
            celda.alignment = Alignment(
                horizontal=(
                    "center"
                    if indice in (1, 3, 4, 5, 6, 11)
                    else "left"
                ),
                vertical="center",
                wrap_text=True,
            )

        # N.º Biopsia muestra Biopsia + Cantidad en una sola línea.
        hoja.row_dimensions[numero_fila].height = 60

    hoja.freeze_panes = "A2"

    if hoja.max_row >= 2:
        hoja.auto_filter.ref = (
            f"A1:{get_column_letter(len(COLUMNAS_REPORTE))}"
            f"{hoja.max_row}"
        )

    anchos = {
        "Fecha": 13,
        "N.º Biopsia": 28,
        "Médico": 22,
        "Enfermera": 22,
        "Técnica": 25,
        "Observaciones": 30,
        "Firma": 18,
    }

    for indice, encabezado in enumerate(
        COLUMNAS_REPORTE,
        start=1,
    ):
        hoja.column_dimensions[
            get_column_letter(indice)
        ].width = anchos[encabezado]

    hoja.row_dimensions[1].height = 45
    hoja.sheet_view.showGridLines = False

    # Para imprimirlo como un registro ancho, similar a la libreta.
    hoja.page_setup.orientation = "landscape"
    hoja.page_setup.fitToWidth = 1
    hoja.page_setup.fitToHeight = 0

    archivo = BytesIO()
    libro.save(archivo)
    archivo.seek(0)

    return archivo.getvalue()


def filas_para_vista_previa(filas):
    """
    Devuelve solo las columnas visibles en el orden del reporte.
    """
    return [
        {
            columna: fila.get(columna, "")
            for columna in COLUMNAS_REPORTE
        }
        for fila in filas
    ]


# =========================================================
# MENÚ LATERAL
# =========================================================

with st.sidebar:
    st.markdown(
        '<h2>Gastroenterologia </h2>',
        unsafe_allow_html=True,
    )

    opcion_menu = st.radio(
        "Menú",
        [
            "🏠 Formulario",
            "📋 Reportes",
            "📥 Indicadores",
        ],
        label_visibility="collapsed",
    )

    st.divider()

    st.caption("Cuenta de Google")
    st.info("Autorización al usar Google")
    st.caption("✅ VERSIÓN APPS SCRIPT - 08/08/2026")

    st.markdown(
        """
        <div class="acento-colores">
            <i class="c-verde"></i>
            <i class="c-azul"></i>
            <i class="c-rojo"></i>
        </div>
        """,
        unsafe_allow_html=True,
    )

# =========================================================
# ESTILOS
# =========================================================

# =========================================================
# FONDO HOSPITAL - SOLO ÁREA PRINCIPAL
# =========================================================
if FONDO_HOSPITAL_BASE64:
    st.markdown(
        f"""
        <style>
        [data-testid="stMain"] {{
            background-color: #AEBCC6 !important;
            background-image:
                linear-gradient(
                    rgba(174, 188, 198, 0.66),
                    rgba(142, 160, 172, 0.66)
                ),
                url("data:image/png;base64,{FONDO_HOSPITAL_BASE64}") !important;

            background-position: center center !important;
            background-size: cover !important;
            background-repeat: no-repeat !important;
            background-attachment: scroll !important;
            min-height: 100vh !important;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )

st.markdown(
    """
    <style>
    @import url("https://fonts.googleapis.com/css2?family=Orbitron:wght@500;600;700&family=Press+Start+2P&display=swap");


    /* =========================================================
       FONDO DEL ÁREA PRINCIPAL
       ========================================================= */

    .stApp,
    [data-testid="stAppViewContainer"] {
        background: linear-gradient(
            180deg,
            #C06A7D 0%,
            #AF596E 100%
        ) !important;
    }

        /* =========================================================
       MENÚ LATERAL
       ========================================================= */

    [data-testid="stSidebar"] {
        background: linear-gradient(
            to right,
            #FFFFFF 0%,
            #FFFFFF 70%,
            #C5D0D6 100%
        ) !important;
        border-right: none !important;
        box-shadow: none !important;
        z-index: 20 !important;
        overflow: visible !important;
    }

    [data-testid="stAppViewContainer"] {
        overflow: visible !important;
    }

    [data-testid="stSidebar"] > div:first-child {
        background: linear-gradient(
            to right,
            #FFFFFF 0%,
            #FFFFFF 70%,
            #C5D0D6 100%
        ) !important;
        overflow: visible !important;
        box-shadow: none !important;
    }

    [data-testid="stSidebarContent"] {
        background: linear-gradient(
            to right,
            #FFFFFF 0%,
            #FFFFFF 70%,
            #C5D0D6 100%
        ) !important;
        box-shadow: none !important;
    }

    [data-testid="stSidebar"] [role="radiogroup"] {
        background: transparent !important;
    }

    [data-testid="stSidebar"] [data-testid="stAlert"] {
        background: #F3F6F8 !important;
        border: 1.5px solid #D5DEE4 !important;
        border-radius: 9px !important;
    }

    [data-testid="stSidebar"] [data-testid="stAlert"] p {
        color: #2F4A5A !important;
    }

    [data-testid="stSidebar"] [data-testid="stCaptionContainer"] p,
    [data-testid="stSidebar"] .stCaption p {
        color: #5A7382 !important;
    }

    /* =========================================================
       CAJAS PRINCIPALES
       ========================================================= */

    .st-key-caja_conexion,
    .st-key-caja_excel {
        background: rgba(54, 78, 96, 0.30) !important;
        border: 1.5px solid #7F9EAF !important;
        border-radius: 19px !important;
        padding: 1.35rem 1.50rem 1.25rem 1.50rem !important;

        box-shadow:
            0 5px 14px rgba(25, 43, 55, 0.20),
            inset 0 1px 0 rgba(255, 255, 255, 0.08) !important;
    }

    .st-key-caja_conexion\:hover,
    .st-key-caja_excel\:hover {
        border-color: #8EADBE !important;

        box-shadow:
            0 7px 17px rgba(25, 43, 55, 0.24),
            inset 0 1px 0 rgba(255, 255, 255, 0.10) !important;
    }


    /* =========================================================
       TÍTULO PRINCIPAL
       ========================================================= */

    .titulo-principal {
        color: #F2D79B !important;
        font-size: 31px !important;
        font-weight: 800 !important;
        line-height: 1.20 !important;
        letter-spacing: 0.20px !important;

        padding-bottom: 7px !important;
        margin-top: -2px !important;
        margin-bottom: 3px !important;

        display: inline-block !important;
        width: fit-content !important;
        max-width: none !important;

        border-bottom: 1.5px solid rgba(225, 236, 242, 0.65) !important;
        text-shadow: 0 2px 4px rgba(28, 45, 57, 0.25) !important;
    }

    .subtitulo-principal,
    .subtitulo-principal p {
        color: #FFFFFF !important;
        font-size: 22px !important;
        font-weight: 600 !important;
        line-height: 1.45 !important;
        letter-spacing: 0.20px !important;
        margin: 0 !important;

        text-shadow:
            0 2px 4px rgba(0, 0, 0, 0.85),
            0 0 8px rgba(0, 0, 0, 0.40),
            0 0 14px rgba(0, 0, 0, 0.22) !important;
    }

    /* =========================================================
       TÍTULOS DE LAS CAJAS
       ========================================================= */

    .st-key-caja_conexion h3,
    .st-key-caja_excel h3 {
        color: #F3C087 !important;
        font-size: 25px !important;
        font-weight: 750 !important;
        letter-spacing: 0.15px !important;
        margin-bottom: 5px !important;
        text-shadow: 0 1px 2px rgba(25, 42, 53, 0.22) !important;
    }

    .st-key-caja_conexion [data-testid="stCaptionContainer"],
    .st-key-caja_excel [data-testid="stCaptionContainer"] {
        color: #D7E1E7 !important;
        font-size: 16px !important;
        line-height: 1.45 !important;
    }


    /* =========================================================
       ETIQUETAS Y CAMPOS
       ========================================================= */

    [data-testid="stMain"] label {
        color: #F1F5F7 !important;
    }

    [data-testid="stVerticalBlockBorderWrapper"] {
        background: rgba(54, 78, 96, 0.22) !important;
        border: 1.5px solid #7F9EAF !important;
        border-radius: 14px !important;
        box-shadow: 0 3px 9px rgba(25, 43, 55, 0.14) !important;
    }

    [data-testid="stVerticalBlockBorderWrapper"] p {
        color: #F1F5F7 !important;
    }


    /* =========================================================
       BOTONES
       ========================================================= */

    div.stButton > button {
        background: #A8BDC9 !important;
        color: #182E3D !important;
        border: 1.5px solid #3F7298 !important;
        border-radius: 14px !important;
        font-weight: 500 !important;
        box-shadow: 0 2px 5px rgba(25, 43, 55, 0.15) !important;
    }

    div.stButton > button\:hover {
        background: #97AFBD !important;
        color: #102531 !important;
        border-color: #2F6389 !important;
        box-shadow: 0 4px 8px rgba(25, 43, 55, 0.22) !important;
    }


    /* =========================================================
       CARGADOR DE ARCHIVO EXCEL
       ========================================================= */

    [data-testid="stFileUploaderDropzone"] {
        background: #55778A !important;
        border: 1.5px solid #506F82 !important;
    }

    [data-testid="stFileUploaderDropzone"] span,
    [data-testid="stFileUploaderDropzone"] small {
        color: #182E3D !important;
    }

    [data-testid="stFileUploaderDropzone"] button {
        background: #A8BDC9 !important;
        color: #182E3D !important;
        border: 1.5px solid #3F7298 !important;
        border-radius: 14px !important;
        font-weight: 500 !important;
        box-shadow: 0 2px 5px rgba(25, 43, 55, 0.15) !important;
    }

    [data-testid="stFileUploaderDropzone"] button\:hover {
        background: #97AFBD !important;
        color: #102531 !important;
        border-color: #2F6389 !important;
        box-shadow: 0 4px 8px rgba(25, 43, 55, 0.22) !important;
    }


    /* =========================================================
       TEXTO Y ESPACIADO GENERAL
       ========================================================= */

    [data-testid="stMain"] {
        color: #EDF2F5 !important;
    }

    /* Spinner - texto de conexión con Google */
    [data-testid="stSpinner"] p {
        font-size: 19px !important;
        font-weight: 600 !important;
        color: #F4F7F9 !important;

        text-shadow:
            0 2px 3px rgba(0, 0, 0, 0.65),
            0 0 8px rgba(0, 0, 0, 0.25) !important;
    }

    [data-testid="stSpinner"] {
        transform: translateY(-40px) !important;
        display: flex !important;
        align-items: center !important;
    }

    /* Ocultar el circulito original de Streamlit */
    [data-testid="stSpinner"] svg {
        display: none !important;
    }

    /* Tres barras verdes en movimiento */
    [data-testid="stSpinner"]::before {
        content: "" !important;
        width: 30px !important;
        height: 24px !important;
        margin-right: 12px !important;
        flex-shrink: 0 !important;

        background:
            linear-gradient(#58D68D 0 0) left bottom / 6px 45% no-repeat,
            linear-gradient(#58D68D 0 0) center bottom / 6px 90% no-repeat,
            linear-gradient(#58D68D 0 0) right bottom / 6px 60% no-repeat !important;

        transform-origin: center bottom !important;
        animation: barras-cargando 0.55s ease-in-out infinite alternate !important;
        filter: drop-shadow(0 0 4px rgba(88, 214, 141, 0.55)) !important;
    }

    @keyframes barras-cargando {
        0% {
            transform: scaleY(0.48);
        }
        100% {
            transform: scaleY(1);
        }
    }

    /* Subir barra de progreso de creación del formulario */
    [data-testid="stProgress"] {
        transform: translateY(-30px) !important;
    }

    /* Texto Creando formulario... */
    [data-testid="stProgress"] p {
        font-size: 19px !important;
        font-weight: 600 !important;
        color: #F4F7F9 !important;

        text-shadow:
            0 2px 3px rgba(0, 0, 0, 0.65),
            0 0 8px rgba(0, 0, 0, 0.25) !important;
    }

    [data-testid="stMainBlockContainer"] {
        padding-top: 1.2rem !important;
    }

    [data-testid="stMainBlockContainer"] {
        padding-top: 0.6rem !important;
    }

    [data-testid="stHeader"] {
        display: none !important;
    }

    [data-testid="stMain"] [data-testid="stAlert"] p {
        color: #18303F !important;
        font-weight: 500 !important;
    }

    [data-testid="stMain"] [data-testid="stAlert"] {
        border: 1.5px solid #506F82 !important;
        border-radius: 10px !important;
        background: rgba(88, 214, 141, 0.28) !important;

        /* Ocupa exactamente el ancho de la primera columna */
        width: 100% !important;
        max-width: 100% !important;

        /* Queda alineada con el botón 1 */
        margin-left: 0 !important;
        margin-right: 0 !important;

        /* Altura */
        padding: 0 !important;
        height: 50px !important;
        min-height: 50px !important;
        max-height: 50px !important;
        box-sizing: border-box !important;

        display: flex !important;
        align-items: center !important;

        box-shadow:
            0 8px 20px rgba(20, 45, 38, 0.30),
            0 3px 7px rgba(0, 0, 0, 0.14),
            inset 0 1px 0 rgba(255, 255, 255, 0.12) !important;

        /* Separación uniforme respecto a los botones */
        transform: translateY(-4px) !important;
    }

    /* Quitar padding interno de Streamlit */
    [data-testid="stMain"] [data-testid="stAlert"] > div {
        padding-top: 0 !important;
        padding-bottom: 0 !important;
        min-height: 0 !important;
        height: 100% !important;

        display: flex !important;
        align-items: center !important;
    }

    /* Texto */
    [data-testid="stMain"] [data-testid="stAlert"] p {
        white-space: nowrap !important;
        margin: 0 !important;
        line-height: 1.1 !important;
    }

    /* Tamaño de todos los títulos principales */
    [data-testid="stMain"] h1 {
        font-size: 32px !important;
        line-height: 1.15 !important;
        margin-bottom: 12px !important;
    }

    .descripcion-seccion {
        color: #F4E3B2 !important;
        font-size: 20px !important;
        font-weight: 600 !important;
        line-height: 1.45 !important;
        margin-top: -18px !important;
        margin-bottom: 20px !important;
        text-shadow: 0 1px 2px rgba(25, 43, 55, 0.30) !important;
    }

    /* =========================================================
    CAMPOS DE LA PLANTILLA
    ========================================================= */

    .campo-etiqueta {
        color: #F1F5F7 !important;
        font-size: 17px !important;
        font-weight: 600 !important;
        line-height: 1.2 !important;
        margin: 0 !important;
        padding: 0 !important;

        /* Sube ligeramente el texto */
        transform: translateY(-11px) !important;
        white-space: nowrap !important;
    }

    .campo-obligatorio {
        color: #FFD18A !important;
        font-size: 19px !important;
        font-weight: 800 !important;
        margin-left: 3px !important;
    }

    .leyenda-obligatorio {
        color: #E9EEF1 !important;
        font-size: 14px !important;
        margin-top: 10px !important;
        margin-bottom: 4px !important;
    }



    /* =========================================================
       AJUSTES MÍNIMOS SOLICITADOS
       - Mantiene el fondo general sin cambios.
       - Aumenta solo el texto de las opciones del menú lateral.
       - Aclara ligeramente solo las cajas de la plantilla.
       ========================================================= */

    /* Texto de las opciones del menú lateral */
    [data-testid="stSidebar"] [role="radiogroup"] label p,
    [data-testid="stSidebar"] [role="radiogroup"] label span {
        font-size: 17px !important;
        font-weight: 500 !important;
        line-height: 1.35 !important;
        color: #2F4A5A !important;
    }

    /* Separación entre INDICE y la lista del menú */
    
        /* Texto de las opciones del menú lateral */
    [data-testid="stSidebar"] [role="radiogroup"] label p,
    [data-testid="stSidebar"] [role="radiogroup"] label span {
        font-size: 17px !important;
        font-weight: 500 !important;
        line-height: 1.35 !important;
        color: #2F4A5A !important;
    }

    [data-testid="stSidebar"] [role="radiogroup"] label {
        border-radius: 9px !important;
        padding: 4px 6px 4px 4px !important;
    }

    [data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked) {
        background: #F3F6F8 !important;
    }

    [data-testid="stSidebar"] h2 {
        margin-top: 1.5rem !important;
        margin-bottom: 0.30rem !important;
        color: #2F4A5A !important;
        font-family: "Press Start 2P", "Courier New", monospace !important;
        font-size: 14px !important;
        font-weight: 400 !important;
        letter-spacing: 0.15px !important;
        line-height: 1.35 !important;
        text-shadow:
            0 2px 3px rgba(0, 0, 0, 0.45),
            0 0 3px rgba(255, 255, 255, 0.08) !important;
        text-decoration: none !important;
        border-bottom: none !important;
        box-shadow: none !important;
        display: inline-block !important;
        position: relative !important;
        padding-left: 0.55rem !important;
    }

    [data-testid="stSidebar"] .indice-icono {
        font-size: 22px !important;
        vertical-align: middle !important;
        margin-right: -2px !important;
    }

    [data-testid="stSidebar"] h2::after {
        content: "" !important;
        position: absolute !important;
        left: 0.55rem !important;
        bottom: -4px !important;
        width: calc(100% - 0.55rem) !important;
        height: 1px !important;
        background: #9AADB8 !important;
        border-radius: 999px !important;
        opacity: 0.85 !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.70),
            0 1px 2px rgba(47, 74, 90, 0.18) !important;
    }

    [data-testid="stSidebar"] hr {
        border: none !important;
        border-top: 1px solid #9AADB8 !important;
        opacity: 0.80 !important;
        box-shadow: 0 1px 0 rgba(255, 255, 255, 0.70) !important;
    }

    [data-testid="stSidebar"] .acento-colores {
    width: auto !important;
    height: auto !important;
    margin: 70px auto 16px auto !important;
    border: none !important;
    background: none !important;
    box-shadow: none !important;
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    gap: 6px !important;
}

    [data-testid="stSidebar"] .acento-colores i {
        display: block !important;
        width: 6px !important;
        height: 6px !important;
        margin: 0 !important;
        padding: 0 !important;
        flex: 0 0 auto !important;
        border-radius: 50% !important;
        border: none !important;
    }

    [data-testid="stSidebar"] .c-verde {
    animation: color-1 3.2s linear infinite !important;
    }

    [data-testid="stSidebar"] .c-azul {
        animation: color-2 3.2s linear infinite !important;
    }

    [data-testid="stSidebar"] .c-rojo {
        animation: color-3 3.2s linear infinite !important;
    }

    @keyframes color-1 {
        0%, 100% { background-color: #27AE60; }
        25%      { background-color: #2E86C1; }
        50%      { background-color: #E74C3C; }
        75%      { background-color: #F1C40F; }
    }

    @keyframes color-2 {
        0%, 100% { background-color: #2E86C1; }
        25%      { background-color: #E74C3C; }
        50%      { background-color: #F1C40F; }
        75%      { background-color: #27AE60; }
    }

    @keyframes color-3 {
        0%, 100% { background-color: #E74C3C; }
        25%      { background-color: #F1C40F; }
        50%      { background-color: #27AE60; }
        75%      { background-color: #2E86C1; }
    }

    [data-testid="stSidebar"] [role="radiogroup"] {
        margin-top: 2.05rem !important;
    }

    /* =========================================================
       TARJETA INICIAL DEL FORMULARIO
       ========================================================= */
    .st-key-tarjeta_inicio_formulario {
        position: relative !important;
        background: rgba(54, 78, 96, 0.28) !important;
        border: 1.8px solid #000000 !important;
        border-radius: 22px !important;
        padding: 2.05rem 2.20rem 2.05rem 2.20rem !important;
        margin-top: 1.2rem !important;
        min-height: 230px !important;
        overflow: hidden !important;
        isolation: isolate !important;
        display: flex !important;
        flex-direction: column !important;
        justify-content: center !important;
        

        box-shadow:
            0 8px 22px rgba(0, 0, 0, 0.18),
            inset 0 1px 0 rgba(255, 255, 255, 0.08) !important;
    }

    /* =========================================================
       DOS LUCES CORTAS QUE RECORREN EL BORDE REAL
       ========================================================= */

    .st-key-tarjeta_inicio_formulario::before,
    .st-key-tarjeta_inicio_formulario::after {
        content: "" !important;
        position: absolute !important;
        left: 0 !important;
        top: 0 !important;

        /* Segmento corto de luz */
        width: 9px !important;
        height: 2.5px !important;
        border-radius: 999px !important;

        background: linear-gradient(
            90deg,
            transparent 0%,
            rgba(205, 229, 56, 0.18) 12%,
            #CDE538 34%,
            #F1FF79 50%,
            #CDE538 66%,
            rgba(205, 229, 56, 0.18) 88%,
            transparent 100%
        ) !important;

        box-shadow:
            0 0 2px rgba(205, 229, 56, 0.45),
            0 0 4px rgba(205, 229, 56, 0.20) !important;

        /* La trayectoria es exactamente el perímetro de la tarjeta */
        offset-path: inset(1px round 20px) !important;
        offset-anchor: 50% 50% !important;
        offset-rotate: auto !important;

        pointer-events: none !important;
        z-index: 6 !important;
        will-change: offset-distance !important;
    }

    .st-key-tarjeta_inicio_formulario [data-testid="stVerticalBlock"],
    .st-key-tarjeta_inicio_formulario [data-testid="stVerticalBlockBorderWrapper"],
    .st-key-tarjeta_inicio_formulario [data-testid="element-container"] {
        height: auto !important;
        min-height: 0 !important;
        flex: 0 1 auto !important;
    }

    .st-key-tarjeta_inicio_formulario::before {
        animation: luz_borde_horaria 8.5s linear infinite !important;
    }

    .st-key-tarjeta_inicio_formulario::after {
        animation: luz_borde_antihoraria 8.5s linear infinite !important;
    }

    @keyframes luz_borde_horaria {
        from {
            offset-distance: 0%;
        }
        to {
            offset-distance: 100%;
        }
    }

    @keyframes luz_borde_antihoraria {
        from {
            offset-distance: 50%;
        }
        to {
            offset-distance: -50%;
        }
    }

    .inicio-titulo {
        color: #B7CBD3 !important;
        font-family: "Press Start 2P", "Courier New", monospace !important;
        font-size: 18px !important;
        font-weight: 400 !important;
        line-height: 1.55 !important;
        letter-spacing: 0.25px !important;
        text-align: center !important;

        margin: 2px auto 12px auto !important;
        width: max-content !important;

       text-shadow:
            -1px -1px 0 rgba(255, 255, 255, 0.35),
            1px 1px 1px rgba(0, 0, 0, 0.55),
            1.5px 2.5px 5px rgba(0, 0, 0, 0.45) !important;

        text-decoration: none !important;
        border-bottom: none !important;
        box-shadow: none !important;

        position: relative !important;
        padding-bottom: 10px !important;
    }

    .inicio-titulo::after {
        content: "" !important;
        position: absolute !important;
        left: 0 !important;
        bottom: 0 !important;
        width: 100% !important;
        height: 2px !important;
        background: #CDE538 !important;
        box-shadow: none !important;
        border-radius: 0 !important;
        display: none !important;
    }

    .inicio-subtitulo {
        color: #F3B562 !important;
        font-family: "Press Start 2P", "Courier New", monospace !important;
        font-size: 13px !important;
        font-weight: 400 !important;
        line-height: 1.6 !important;

        letter-spacing: 0.20px !important;
        text-align: center !important;
        margin-bottom: 18px !important;
        text-shadow: 0 2px 4px rgba(25, 43, 55, 0.24) !important;
        text-decoration: none !important;
        border-bottom: none !important;
        box-shadow: none !important;
        display: inline-flex !important;
        align-items: center !important;
        justify-content: center !important;
        gap: 10px !important;
        width: 100% !important;
    }

    .inicio-subtitulo::before {
        content: "" !important;
        display: inline-block !important;
        width: 4px !important;
        height: 20px !important;
        border-radius: 999px !important;
        background: #CDE538 !important;
        flex: 0 0 auto !important;
    }

    .inicio-descripcion {
        color: #FFFFFF !important;
        font-size: 18px !important;
        font-weight: 600 !important;
        line-height: 1.45 !important;
        letter-spacing: 0.10px !important;
        text-align: center !important;
        margin-bottom: 38px !important;

        text-shadow:
            0 1px 2px rgba(0, 0, 0, 0.55),
            0 0 4px rgba(0, 0, 0, 0.18) !important;
    }

    .inicio-estado-formulario {
        width: fit-content !important;
        margin: -12px auto 18px auto !important;
        padding: 6px 14px !important;

        color: #FFF59A !important;
        background: transparent !important;
        border: none !important;
        border-radius: 0 !important;
        box-shadow: none !important;

        font-size: 17px !important;
        font-weight: 600 !important;
        letter-spacing: 0.15px !important;
        text-align: center !important;

        text-shadow:
            0 2px 4px rgba(0, 0, 0, 0.62),
            0 0 5px rgba(0, 0, 0, 0.22) !important;

    }

    .inicio-icono {
        text-align: center !important;
        font-size: 46px !important;
        line-height: 1 !important;
        margin-bottom: 4px !important;
    }

    /* Logo principal: más grande y sin marco visual adicional */
    .st-key-tarjeta_inicio_formulario [data-testid="stImage"] img {
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
        margin-top: 0 !important;
    }

    /* =========================================================
       FLUJO PRINCIPAL DE 3 PASOS
       ========================================================= */

    .st-key-flujo_paso_1 div.stButton > button,
    .st-key-flujo_paso_2 div.stButton > button,
    .st-key-flujo_paso_3 div.stButton > button,
    .st-key-flujo_paso_2 [data-testid="stLinkButton"] a,
    .st-key-flujo_paso_3 [data-testid="stLinkButton"] a {
        width: 100% !important;
        min-height: 52px !important;
        height: 52px !important;
        padding: 0.48rem 0.70rem !important;

        background: linear-gradient(
            180deg,
            #CC7488 0%,
            #BC6579 100%
        ) !important;

        color: #FFFFFF !important;

        border: 1.5px solid #733547 !important;

        border-radius: 13px !important;

        font-family: "Orbitron", "Segoe UI", sans-serif !important;
        font-size: 14px !important;
        font-weight: 700 !important;
        letter-spacing: 0.45px !important;

        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        white-space: nowrap !important;

        box-shadow:
            0 5px 12px rgba(25, 43, 55, 0.30),
            0 2px 5px rgba(0, 0, 0, 0.15) !important;
    }

    .st-key-flujo_paso_1 div.stButton > button:hover,
    .st-key-flujo_paso_2 [data-testid="stLinkButton"] a:hover,
    .st-key-flujo_paso_3 [data-testid="stLinkButton"] a:hover {

        background: linear-gradient(
            180deg,
            #B95A72 0%,
            #A44761 100%
        ) !important;

        border-color: #7D3147 !important;

        box-shadow:
            0 0 10px rgba(164,71,97,0.30),
            0 6px 14px rgba(20,35,45,0.20) !important;
    }

    .st-key-flujo_paso_1 div.stButton > button p,
    .st-key-flujo_paso_1 div.stButton > button span,
    .st-key-flujo_paso_2 div.stButton > button p,
    .st-key-flujo_paso_2 div.stButton > button span,
    .st-key-flujo_paso_3 div.stButton > button p,
    .st-key-flujo_paso_3 div.stButton > button span,
    .st-key-flujo_paso_2 [data-testid="stLinkButton"] a p,
    .st-key-flujo_paso_2 [data-testid="stLinkButton"] a span,
    .st-key-flujo_paso_3 [data-testid="stLinkButton"] a p,
    .st-key-flujo_paso_3 [data-testid="stLinkButton"] a span {
        color: #F2F5F7 !important;
        font-family: "Orbitron", "Segoe UI", sans-serif !important;
        font-size: 14px !important;
        font-weight: 700 !important;
        letter-spacing: 0.45px !important;
        margin: 0 !important;
    }

    .st-key-flujo_paso_1 div.stButton > button:disabled,
    .st-key-flujo_paso_2 div.stButton > button:disabled,
    .st-key-flujo_paso_3 div.stButton > button:disabled {

        background: linear-gradient(
            180deg,
            rgba(197,109,128,0.65) 0%,
            rgba(181,93,113,0.65) 100%
        ) !important;

        color: rgba(255,255,255,0.90) !important;

        border: none !important;

        box-shadow:
            0 5px 12px rgba(25,43,55,0.18),
            0 2px 5px rgba(0,0,0,0.08) !important;

        cursor: not-allowed !important;
    }
    
    /* Al pasar el mouse por botones desactivados */
        .st-key-flujo_paso_1 div.stButton > button:disabled:hover,
        .st-key-flujo_paso_2 div.stButton > button:disabled:hover,
        .st-key-flujo_paso_3 div.stButton > button:disabled:hover {
            border-color: #66727A !important;
        }

    /* =========================================================
       INDICADOR VISUAL DEL PASO 2
       ========================================================= */

    .indicador-paso-2 {
        width: 100% !important;
        min-height: 58px !important;

        display: flex !important;
        flex-direction: column !important;
        align-items: center !important;
        justify-content: flex-start !important;

        text-align: center !important;
        margin: 0 !important;
        padding: 0 !important;

        /* Centrado debajo del segundo botón */
        transform: translateY(-9px) !important;
    }

    .triangulo-paso-2 {
        color: #FF8C00 !important;
        font-size: 36px !important;
        line-height: 0.78 !important;
        font-weight: 900 !important;

        margin: 0 !important;
        padding: 0 !important;

        text-shadow:
            0 0 5px rgba(255, 140, 0, 0.95),
            0 0 11px rgba(255, 140, 0, 0.62),
            0 3px 4px rgba(0, 0, 0, 0.35) !important;

        filter: drop-shadow(0 0 4px rgba(255, 140, 0, 0.45)) !important;
        animation: paso2-indicador 0.72s ease-in-out infinite alternate !important;
    }

    .texto-paso-2 {
        color: #FFD18A !important;
        font-family: "Orbitron", "Segoe UI", sans-serif !important;
        font-size: 12px !important;
        font-weight: 700 !important;
        letter-spacing: 1.25px !important;

        margin-top: 7px !important;
        line-height: 1 !important;

        text-shadow:
            0 2px 3px rgba(0, 0, 0, 0.60),
            0 0 6px rgba(255, 140, 0, 0.22) !important;
    }

    @keyframes paso2-indicador {
        from {
            transform: translateY(4px) scale(1);
        }
        to {
            transform: translateY(-4px) scale(1.06);
        }
    }

    /* Ver/Ocultar plantilla:
       fuera de la tarjeta y pegado a la esquina inferior derecha */
    .st-key-boton_ver_plantilla {
        position: relative !important;
        transform: translate(-30px, -48px) !important;
        margin-bottom: -46px !important;
        z-index: 20 !important;
    }

    .st-key-boton_ver_plantilla div.stButton {
        display: flex !important;
        justify-content: flex-start !important;
    }

    .st-key-boton_ver_plantilla div.stButton > button {
        width: auto !important;
        min-width: 0 !important;
        min-height: 42px !important;
        padding: 0.65rem 1.3rem !important;
        white-space: nowrap !important;

        background: rgba(70, 95, 110, 0.38) !important;
        border: 1px solid rgba(225, 235, 240, 0.42) !important;
        border-radius: 9px !important;

        box-shadow: 0 2px 5px rgba(0, 0, 0, 0.12) !important;

        color: #F5F7F8 !important;
        font-size: 14px !important;
        font-weight: 500 !important;
        text-shadow: 0 2px 4px rgba(0, 0, 0, 0.55) !important;
    }

    .st-key-boton_ver_plantilla,
    .st-key-boton_ver_plantilla div.stButton,
    .st-key-boton_ver_plantilla div.stButton > button {
        flex-shrink: 0 !important;
    }

    .st-key-boton_ver_plantilla div.stButton > button:hover {
        background: rgba(70, 95, 110, 0.52) !important;
        border-color: rgba(235, 242, 245, 0.60) !important;
        color: #FFFFFF !important;

        box-shadow: 0 3px 7px rgba(0, 0, 0, 0.17) !important;
    }

    /* Solo las cajas de cada campo de la Plantilla de formulario */
    div[class*="st-key-vista_campo_"] {
        background: rgba(190, 207, 217, 0.32) !important;
        border: 1.5px solid #9DB2BE !important;
        border-radius: 14px !important;
        box-shadow: 0 3px 9px rgba(25, 43, 55, 0.10) !important;
    }

    .footer-informatica {
    position: fixed !important;
    bottom: 18px !important;
    left: calc(50% + 10.5rem) !important;
    transform: translateX(-50%) !important;

    color: rgba(230,230,230,0.95) !important;
    font-size: 15px !important;      /* más grande */
    font-weight: 600 !important;
    letter-spacing: 0.6px !important;

    background: transparent !important;
    border: none !important;
    box-shadow: none !important;

    text-shadow:
        0 0 3px rgba(0,0,0,1),
        0 0 6px rgba(0,0,0,0.95),
        0 0 12px rgba(0,0,0,0.90),
        2px 2px 4px rgba(0,0,0,0.90) !important;
}

    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
        """
        <div class="acento-colores"></div>
        """,
        unsafe_allow_html=True,
    )

st.markdown(
    """
    <div class="footer-informatica">
        OGEI - Hospital San Juan de Matucana
    </div>
    """,
    unsafe_allow_html=True,
)

st.markdown("""
<style>
.block-container{
    padding-top: 0.5rem !important;
}
</style>
""", unsafe_allow_html=True)

# =========================================================
# CONTENIDO SEGÚN EL MENÚ
# =========================================================

if opcion_menu == "🏠 Formulario":

    formulario_hoy = None
    error_consulta_formulario_hoy = None

    fecha_hoy_sesion = datetime.now(
        ZoneInfo("America/Lima")
    ).strftime("%Y-%m-%d")

    try:
        formulario_hoy = obtener_formulario_hoy()

        if formulario_hoy:
            cargar_formulario_hoy_en_sesion(formulario_hoy)
        else:
            # Si cambió el día, no se debe seguir usando el formulario anterior.
            st.session_state.pop("form_url_base", None)
            st.session_state.pop("form_entry_id_fecha", None)
            st.session_state.pop("form_url_editar", None)
            st.session_state.pop("form_fecha_activa", None)

    except requests.exceptions.RequestException as error:
        error_consulta_formulario_hoy = error

        # Si no podemos verificar el backend, solo conservamos un formulario
        # que sepamos que pertenece al día actual.
        if st.session_state.get("form_fecha_activa") != fecha_hoy_sesion:
            st.session_state.pop("form_url_base", None)
            st.session_state.pop("form_entry_id_fecha", None)
            st.session_state.pop("form_url_editar", None)
            st.session_state.pop("form_fecha_activa", None)

    if "mostrar_plantilla" not in st.session_state:
        st.session_state.mostrar_plantilla = False

    texto_boton_plantilla = (
        "🙈 Ocultar plantilla"
        if st.session_state.mostrar_plantilla
        else "👁️ Ver plantilla"
    )

    # =========================================================
    # TARJETA INICIAL
    # =========================================================

    
    logo_esquina_izq_b64 = obtener_imagen_base64(RUTA_LOGO_DERECHA)
    logo_esquina_der_b64 = obtener_imagen_base64(RUTA_LOGO_IZQUIERDA)

    st.markdown(
        f"""
        <style>
        .logo-esquina {{
            position: fixed !important;
            top: 14px !important;
            z-index: 15 !important;
            pointer-events: none !important;
        }}

        .logo-esquina img {{
            width: 100px !important;
            height: auto !important;
            display: block !important;
            filter: drop-shadow(0 2px 5px rgba(0, 0, 0, 0.25));
        }}

        .logo-esquina-izquierda {{
            left: calc(21rem + 16px) !important;
        }}

        .logo-esquina-derecha {{
            right: 40px !important;
        }}

        </style>

        <div class="logo-esquina logo-esquina-izquierda">
            <img src="data:image/png;base64,{logo_esquina_izq_b64}" />
        </div>
        <div class="logo-esquina logo-esquina-derecha">
            <img src="data:image/png;base64,{logo_esquina_der_b64}" />
        </div>
        """,
        unsafe_allow_html=True,
    )

    margen_izq, columna_tarjeta, margen_der = st.columns(
        [1, 7.2, 1]
    )

    with columna_tarjeta:

    

        with st.container(
                            border=True,
                            key="tarjeta_inicio_formulario",
                        ):
                                
            if RUTA_LOGO.exists():
                logo_transparente = obtener_logo_transparente(RUTA_LOGO)

                logo_izq, logo_centro, logo_der = st.columns(
                    [4.15, 1.70, 4.15]
                )

                with logo_centro:
                    st.image(
                        logo_transparente,
                        width=125,
                    )
            else:
                st.markdown(
                    '<div class="inicio-icono">🩺</div>',
                    unsafe_allow_html=True,
                )

            fecha_estado = datetime.now(
                ZoneInfo("America/Lima")
            ).strftime("%d/%m/%Y")

            estado_formulario = ""

            if formulario_hoy:
                estado_formulario = (
                    f'<div class="inicio-estado-formulario">'
                    f'✓ Formulario diario activo · {fecha_estado}'
                    f'</div>'
                )

            st.markdown(
                f"""
                <div class="inicio-titulo">
                    Registro de Biopsias
                </div>

                <div class="inicio-descripcion">
                    Gestión de formularios, reportes e indicadores de biopsias.
                </div>

                {estado_formulario}
                """,
                unsafe_allow_html=True,
            )

    # Ver plantilla va en una columna lateral propia,
    # fuera de la tarjeta y junto a su esquina inferior derecha.
    ojo_margen_izq, ojo_centro, ojo_costado = st.columns(
        [1.4, 7.8, 0.8]
    )

    with ojo_costado:
        boton_ver_plantilla = st.button(
            texto_boton_plantilla,
            key="boton_ver_plantilla",
        )

    # Flujo principal: los 3 pasos quedan en una sola fila
    # con el mismo ancho de la tarjeta principal.
    flujo_margen_izq, columna_flujo, flujo_margen_der = st.columns(
        [1.4, 7.2, 1.4]
    )

    with columna_flujo:
        col_paso1, col_paso2, col_paso3 = st.columns(3)

        with col_paso1:
            with st.container(key="flujo_paso_1"):
                if formulario_hoy:
                    boton_iniciar_formulario = st.button(
                        "✓  1. Formulario de hoy",
                        type="primary",
                        use_container_width=True,
                        disabled=True,
                        key="boton_crear_formulario",
                    )
                else:
                    boton_iniciar_formulario = st.button(
                        "▶  1. Crear formulario",
                        type="primary",
                        use_container_width=True,
                        disabled=(error_consulta_formulario_hoy is not None),
                        key="boton_crear_formulario",
                    )

    if boton_ver_plantilla:
        st.session_state.mostrar_plantilla = (
            not st.session_state.mostrar_plantilla
        )
        st.rerun()

    if error_consulta_formulario_hoy is not None:
        st.warning(
            "No se pudo verificar si ya existe el formulario de hoy. "
            "Por seguridad, la creación queda temporalmente deshabilitada."
        )

    st.write("")

    # =========================================================
    # CAMPOS DEL FORMULARIO EN FILAS COMPACTAS
    # =========================================================

    if st.session_state.mostrar_plantilla:
        for numero, campo in enumerate(
            CAMPOS_FORMULARIO,
            start=1,
        ):
            titulo_campo = campo["titulo"]
            tipo_campo = campo["tipo"]
            obligatorio = campo.get("obligatorio", False)

            with st.container(
                border=True,
                key=f"vista_campo_{numero}",
            ):
                columna_etiqueta, columna_entrada = st.columns(
                    [2.4, 7.6],
                    vertical_alignment="center",
                )

                with columna_etiqueta:
                    marca_obligatorio = (
                        '<span class="campo-obligatorio">*</span>'
                        if obligatorio
                        else ""
                    )

                    st.markdown(
                        f"""
                        <div class="campo-etiqueta">
                            {numero}. {titulo_campo}
                            {marca_obligatorio}
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                with columna_entrada:
                    if tipo_campo == "fecha":
                        st.date_input(
                            titulo_campo,
                            value=date.today(),
                            disabled=False,
                            label_visibility="collapsed",
                            key=f"campo_fecha_{numero}",
                        )

                    elif tipo_campo == "check":
                        st.checkbox(
                            "Sí",
                            value=False,
                            disabled=True,
                            key=f"campo_check_{numero}",
                        )

                    elif tipo_campo == "texto":
                        st.text_input(
                            titulo_campo,
                            placeholder="Respuesta corta",
                            disabled=True,
                            label_visibility="collapsed",
                            key=f"campo_texto_{numero}",
                        )

                    elif tipo_campo == "parrafo":
                        st.text_area(
                            titulo_campo,
                            placeholder="Respuesta",
                            disabled=True,
                            label_visibility="collapsed",
                            height=75,
                            key=f"campo_parrafo_{numero}",
                        )

                    elif tipo_campo == "biopsia":
                        # Vista previa equivalente a las cuadrículas del
                        # Google Form: opción a la izquierda y cantidad al lado.
                        procedimientos = campo.get(
                            "procedimientos",
                            PROCEDIMIENTOS_BIOPSIA,
                        )
                        categorias = campo.get(
                            "categorias",
                            CATEGORIAS_BIOPSIA,
                        )
                        adicionales = campo.get(
                            "adicionales",
                            PROCEDIMIENTOS_ADICIONALES,
                        )
                        cantidades = campo.get(
                            "cantidades",
                            CANTIDADES_GRID,
                        )

                        (
                            columna_procedimiento,
                            columna_biopsia,
                            columna_adicional,
                        ) = st.columns(
                            [1, 1, 1],
                            gap="large",
                            vertical_alignment="top",
                        )

                        procedimiento_es_cuadricula = campo.get(
                            "procedimiento_cuadricula",
                            True,
                        )

                        with columna_procedimiento:
                            if procedimiento_es_cuadricula:
                                st.markdown("**Procedimiento y cantidad**")
                                encabezado_nombre, encabezado_cantidad = (
                                    st.columns(
                                        [2.1, 1],
                                        vertical_alignment="center",
                                    )
                                )
                                with encabezado_nombre:
                                    st.caption("Procedimiento")
                                with encabezado_cantidad:
                                    st.caption("Cantidad")

                                for indice_opcion, opcion_procedimiento in (
                                    enumerate(procedimientos, start=1)
                                ):
                                    col_opcion, col_cantidad = st.columns(
                                        [2.1, 1],
                                        vertical_alignment="center",
                                    )
                                    with col_opcion:
                                        st.write(opcion_procedimiento)
                                    with col_cantidad:
                                        st.selectbox(
                                            f"Cantidad {opcion_procedimiento}",
                                            options=[""] + list(cantidades),
                                            index=0,
                                            label_visibility="collapsed",
                                            key=(
                                                "cantidad_procedimiento_"
                                                f"{numero}_{indice_opcion}"
                                            ),
                                        )

                                st.text_input(
                                    "Otro procedimiento",
                                    placeholder=(
                                        "Complete solo si usó la fila Otros"
                                    ),
                                    label_visibility="collapsed",
                                    key=f"campo_otro_procedimiento_{numero}",
                                )
                            else:
                                # Procedimiento como selección única
                                # (opción múltiple) + cantidad aparte.
                                st.markdown("**Procedimiento**")
                                st.selectbox(
                                    "Procedimiento",
                                    options=list(procedimientos),
                                    index=0,
                                    label_visibility="collapsed",
                                    key=(
                                        f"campo_procedimiento_unico_{numero}"
                                    ),
                                )
                                st.text_input(
                                    "Otro procedimiento",
                                    placeholder=(
                                        "Complete solo si eligió Otros"
                                    ),
                                    label_visibility="collapsed",
                                    key=f"campo_otro_procedimiento_{numero}",
                                )
                                st.markdown("**Cantidad**")
                                st.text_input(
                                    "Cantidad",
                                    placeholder="Cantidad",
                                    label_visibility="collapsed",
                                    key=(
                                        "campo_cantidad_procedimiento_"
                                        f"{numero}"
                                    ),
                                )

                        with columna_biopsia:
                            st.markdown("**Biopsia y cantidad**")
                            encabezado_nombre, encabezado_cantidad = st.columns(
                                [2.1, 1],
                                vertical_alignment="center",
                            )
                            with encabezado_nombre:
                                st.caption("Biopsia")
                            with encabezado_cantidad:
                                st.caption("Cantidad")

                            for indice_opcion, opcion_biopsia in enumerate(
                                categorias,
                                start=1,
                            ):
                                col_opcion, col_cantidad = st.columns(
                                    [2.1, 1],
                                    vertical_alignment="center",
                                )
                                with col_opcion:
                                    st.write(opcion_biopsia)
                                with col_cantidad:
                                    st.selectbox(
                                        f"Cantidad {opcion_biopsia}",
                                        options=[""] + list(cantidades),
                                        index=0,
                                        label_visibility="collapsed",
                                        key=(
                                            f"cantidad_biopsia_{numero}_"
                                            f"{indice_opcion}"
                                        ),
                                    )

                            st.text_input(
                                "Otra biopsia",
                                placeholder="Complete solo si usó la fila Otros",
                                label_visibility="collapsed",
                                key=f"campo_otra_biopsia_{numero}",
                            )

                        with columna_adicional:
                            st.markdown("**Procedimientos adicionales**")
                            encabezado_nombre, encabezado_cantidad = (
                                st.columns(
                                    [2.1, 1],
                                    vertical_alignment="center",
                                )
                            )
                            with encabezado_nombre:
                                st.caption("Técnica")
                            with encabezado_cantidad:
                                st.caption("Cantidad")

                            for indice_opcion, opcion_adicional in (
                                enumerate(adicionales, start=1)
                            ):
                                col_opcion, col_cantidad = st.columns(
                                    [2.1, 1],
                                    vertical_alignment="center",
                                )
                                with col_opcion:
                                    st.write(opcion_adicional)
                                with col_cantidad:
                                    st.selectbox(
                                        f"Cantidad {opcion_adicional}",
                                        options=[""] + list(cantidades),
                                        index=0,
                                        label_visibility="collapsed",
                                        key=(
                                            "cantidad_adicional_"
                                            f"{numero}_{indice_opcion}"
                                        ),
                                    )

                    elif tipo_campo == "lista":
                        st.selectbox(
                            titulo_campo,
                            options=campo.get("opciones", []),
                            disabled=True,
                            label_visibility="collapsed",
                            key=f"campo_lista_{numero}",
                        )

                    elif tipo_campo == "numero":
                        st.number_input(
                            titulo_campo,
                            min_value=1,
                            value=1,
                            step=1,
                            disabled=True,
                            label_visibility="collapsed",
                            key=f"campo_numero_{numero}",
                        )

        st.info(
            "En N.º Biopsia: Procedimiento es una sola opción con su "
            "cantidad aparte. Biopsia y Procedimientos adicionales (APL, "
            "ELVE, enema, inyectoterapia, clip, polipectomías, "
            "mucosectomía) se muestran como cuadrícula: cada opción en "
            "una fila, cantidad al costado. Deje en blanco lo que no "
            "corresponda. Si usa Otros, complete el texto correspondiente. "
            "La columna Firma no se solicita en Google Forms y queda en "
            "blanco."
        )

        st.markdown(
            """
            <div class="leyenda-obligatorio">
                * Campo obligatorio
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.write("")

    # =========================================================
    # CREAR EL GOOGLE FORM DESDE EL BOTÓN DE LA TARJETA
    # =========================================================

    if boton_iniciar_formulario:
        try:

            # Segunda verificación justo antes de crear. Evita duplicados
            # si otra PC creó el formulario mientras esta pantalla estaba abierta.
            formulario_existente = obtener_formulario_hoy()

            if formulario_existente:
                cargar_formulario_hoy_en_sesion(
                    formulario_existente
                )
                st.info(
                    "El formulario de hoy ya existe. Se utilizará ese mismo formulario."
                )
                st.rerun()

            respuesta_sqlite = None
            error_sqlite = None

            # El spinner permanece activo durante TODO el proceso real:
            # creación del Google Form + registro en SQLite.
            progreso_formulario = st.progress(
                10,
                text="Creando formulario..."
            )

            datos_formulario = crear_google_form(
                CAMPOS_FORMULARIO,
                progreso_formulario,
            )

            # =====================================================
            # GUARDAR EL FORMULARIO EN SQLITE MEDIANTE FASTAPI
            # =====================================================
            try:

                progreso_formulario.progress(
                    95,
                    text="Guardando formulario..."
                )

                respuesta_sqlite = requests.post(
                    f"{URL_BACKEND}/formularios",
                    json={
                        "form_id_google": datos_formulario["form_id"],
                        "titulo": datos_formulario["titulo"],
                        "url_responder": datos_formulario[
                            "enlace_respuestas"
                        ],
                        "url_editar": datos_formulario[
                            "enlace_edicion"
                        ],
                    },
                    timeout=15,
                )

            except requests.exceptions.ConnectionError:
                error_sqlite = "conexion"

            except requests.exceptions.Timeout:
                error_sqlite = "timeout"

            except requests.exceptions.RequestException as error:
                error_sqlite = ("request", error)

            # Al salir del spinner, el proceso terminó realmente.
            # Recién aquí se muestra la caja verde / advertencia / error.
            if respuesta_sqlite is not None:

                if respuesta_sqlite.status_code == 200:

                    progreso_formulario.progress(
                        100,
                        text="Formulario creado"
                    )

                    time.sleep(0.3)

                    progreso_formulario.empty()

                    st.session_state["mostrar_exito_formulario"] = True
                    st.rerun()

                    # Misma alineación horizontal que la fila de los 3 botones
                    margen_estado_izq, columna_estado, margen_estado_der = st.columns(
                        [1.4, 7.2, 1.4]
                    )

                    with columna_estado:

                        col_exito, col_indicador, col_vacio = st.columns(3)

                        # Debajo del botón 1
                        with col_exito:
                            st.success("✅ Formulario creado.")

                        # Debajo del botón 2
                        with col_indicador:
                            st.markdown(
                                """
                                <div class="indicador-paso-2">
                                    <div class="triangulo-paso-2">▲</div>
                                    <div class="texto-paso-2">SIGUIENTE</div>
                                </div>
                                """,
                                unsafe_allow_html=True,
                            )

                elif respuesta_sqlite.status_code == 409:
                    try:
                        detalle_409 = respuesta_sqlite.json().get(
                            "detail",
                            respuesta_sqlite.text,
                        )
                    except ValueError:
                        detalle_409 = respuesta_sqlite.text

                    st.warning(
                        "El Google Form fue creado, pero ya estaba "
                        f"registrado en SQLite: {detalle_409}"
                    )

                else:
                    try:
                        detalle_error = respuesta_sqlite.json().get(
                            "detail",
                            respuesta_sqlite.text,
                        )
                    except ValueError:
                        detalle_error = respuesta_sqlite.text

                    st.error(
                        "El Google Form fue creado, pero no se pudo "
                        f"guardar en SQLite: {detalle_error}"
                    )

            elif error_sqlite == "conexion":
                st.error(
                    "El Google Form fue creado, pero no se pudo "
                    "conectar con FastAPI para guardarlo en SQLite."
                )

            elif error_sqlite == "timeout":
                st.error(
                    "El Google Form fue creado, pero FastAPI tardó "
                    "demasiado en guardarlo en SQLite."
                )

            elif isinstance(error_sqlite, tuple) and error_sqlite[0] == "request":
                st.error(
                    "El Google Form fue creado, pero ocurrió un error "
                    "al guardarlo en SQLite."
                )
                st.code(str(error_sqlite[1]))

            st.session_state["form_url_base"] = (
                datos_formulario["enlace_respuestas"].split("?")[0]
            )

            st.session_state["form_entry_id_fecha"] = (
                datos_formulario["entry_id_fecha"]
            )

            st.session_state["form_url_editar"] = (
                datos_formulario["enlace_edicion"]
            )

            st.session_state["form_fecha_activa"] = datetime.now(
                ZoneInfo("America/Lima")
            ).strftime("%Y-%m-%d")
           
        except FileNotFoundError as error:
            st.error(str(error))

        except HttpError as error:
            st.error(
                    "Google rechazó una solicitud de su API."
            )

            detalle = str(error)

            if getattr(error, "content", None):
                    try:
                        detalle = error.content.decode(
                            "utf-8",
                            errors="replace",
                        )
                    except Exception:
                        pass

            st.code(detalle)

        except Exception as error:
            st.error(
                    "No se pudo crear el Google Form."
            )
            st.code(str(error))


    # =========================================================
    # PASOS 2 Y 3 DEL FLUJO PRINCIPAL
    # =========================================================

    if st.session_state.get("form_url_base"):

        url_base = st.session_state["form_url_base"]

        entry_id_fecha = st.session_state.get(
            "form_entry_id_fecha",
            ""
        )

        fecha_hoy = datetime.now(
            ZoneInfo("America/Lima")
        ).strftime("%Y-%m-%d")

        if entry_id_fecha:
            parametros = urlencode({
                "usp": "pp_url",
                f"entry.{entry_id_fecha}": fecha_hoy,
            })

            url_responder = f"{url_base}?{parametros}"
        else:
            url_responder = url_base

        url_editar = st.session_state.get(
            "form_url_editar",
            ""
        )

        with col_paso2:
            with st.container(key="flujo_paso_2"):
                st.link_button(
                    "2. Registrar respuesta",
                    url_responder,
                    use_container_width=True,
                )

        with col_paso3:
            with st.container(key="flujo_paso_3"):
                if url_editar:
                    url_respuestas = url_editar.split("#")[0] + "#responses"

                    st.link_button(
                        "3. Ver respuestas",
                        url_respuestas,
                        use_container_width=True,
                    )
                else:
                    st.button(
                        "3. Ver respuestas",
                        disabled=True,
                        use_container_width=True,
                        key="paso_ver_disabled",
                    )

    else:
        with col_paso2:
            with st.container(key="flujo_paso_2"):
                st.button(
                    "2. Registrar respuesta",
                    disabled=True,
                    use_container_width=True,
                    key="paso_registrar_disabled",
                )

        with col_paso3:
            with st.container(key="flujo_paso_3"):
                st.button(
                    "3. Ver respuestas",
                    disabled=True,
                    use_container_width=True,
                    key="paso_ver_disabled",
                )


elif opcion_menu == "📋 Reportes":
    st.title("📋 Reportes")

    st.markdown(
        """
        <p class="descripcion-seccion">
            Descarga el registro de biopsias por día o por mes.
        </p>
        """,
        unsafe_allow_html=True,
    )

    try:
        respuesta_formularios = requests.get(
            f"{URL_BACKEND}/formularios",
            timeout=10,
        )

        respuesta_formularios.raise_for_status()

        formularios = respuesta_formularios.json().get(
            "formularios",
            [],
        )

        if not formularios:
            st.info(
                "Todavía no hay formularios guardados."
            )

        else:
            indice_formulario = st.selectbox(
                "Seleccionar formulario",
                options=range(len(formularios)),
                format_func=lambda indice: (
                    f"{formularios[indice].get('titulo', 'Formulario')} "
                    f"— {formularios[indice].get('fecha_creacion', '')}"
                ),
                key="formulario_reporte_excel",
            )

            formulario = formularios[indice_formulario]
            form_id = formulario["form_id_google"]

            tipo_reporte = st.radio(
                "Tipo de reporte",
                options=[
                    "Por día",
                    "Por mes",
                ],
                horizontal=True,
                key="tipo_reporte_excel",
            )

            with st.spinner(
                "Consultando respuestas de Google Forms..."
            ):
                registros = obtener_registros_formulario(
                    form_id
                )

            registros_con_fecha = [
                registro
                for registro in registros
                if registro.get("_fecha") is not None
            ]

            if not registros_con_fecha:
                st.info(
                    "El formulario todavía no tiene respuestas "
                    "con una fecha válida."
                )

            else:
                fechas_disponibles = sorted(
                    {
                        registro["_fecha"]
                        for registro in registros_con_fecha
                    }
                )

                if tipo_reporte == "Por día":
                    fecha_predeterminada = fechas_disponibles[-1]

                    fecha_seleccionada = st.date_input(
                        "Fecha del reporte",
                        value=fecha_predeterminada,
                        min_value=fechas_disponibles[0],
                        max_value=fechas_disponibles[-1],
                        key="fecha_reporte_dia",
                    )

                    registros_filtrados = [
                        registro
                        for registro in registros_con_fecha
                        if registro["_fecha"]
                        == fecha_seleccionada
                    ]

                    nombre_hoja = "Registro diario"
                    nombre_archivo = (
                        "registro_biopsias_"
                        f"{fecha_seleccionada:%Y-%m-%d}.xlsx"
                    )

                    periodo_texto = (
                        fecha_seleccionada.strftime("%d/%m/%Y")
                    )

                else:
                    anios_disponibles = sorted(
                        {
                            registro["_fecha"].year
                            for registro in registros_con_fecha
                        },
                        reverse=True,
                    )

                    columna_anio, columna_mes = st.columns(2)

                    with columna_anio:
                        anio_seleccionado = st.selectbox(
                            "Año",
                            options=anios_disponibles,
                            key="anio_reporte_mes",
                        )

                    meses_disponibles = sorted(
                        {
                            registro["_fecha"].month
                            for registro in registros_con_fecha
                            if registro["_fecha"].year
                            == anio_seleccionado
                        }
                    )

                    with columna_mes:
                        mes_seleccionado = st.selectbox(
                            "Mes",
                            options=meses_disponibles,
                            format_func=lambda mes: MESES_ES[mes],
                            key="mes_reporte_mes",
                        )

                    registros_filtrados = [
                        registro
                        for registro in registros_con_fecha
                        if (
                            registro["_fecha"].year
                            == anio_seleccionado
                            and registro["_fecha"].month
                            == mes_seleccionado
                        )
                    ]

                    nombre_hoja = "Registro mensual"
                    nombre_archivo = (
                        "registro_biopsias_"
                        f"{anio_seleccionado}-"
                        f"{mes_seleccionado:02d}.xlsx"
                    )

                    periodo_texto = (
                        f"{MESES_ES[mes_seleccionado]} "
                        f"{anio_seleccionado}"
                    )

                if not registros_filtrados:
                    st.warning(
                        "No existen respuestas en el periodo "
                        "seleccionado."
                    )

                else:
                    filas_reporte = construir_filas_reporte(
                        registros_filtrados
                    )

                    st.metric(
                        "Registros",
                        len(filas_reporte),
                    )

                    st.caption(
                        f"Periodo seleccionado: {periodo_texto}"
                    )

                    contenido_excel = crear_excel_registro(
                        filas_reporte,
                        nombre_hoja,
                    )

                    st.download_button(
                        "⬇️ Descargar registro en Excel",
                        data=contenido_excel,
                        file_name=nombre_archivo,
                        mime=(
                            "application/vnd.openxmlformats-"
                            "officedocument.spreadsheetml.sheet"
                        ),
                        use_container_width=True,
                        type="primary",
                        key="descargar_reporte_excel",
                    )

                    st.subheader("Vista previa")

                    st.dataframe(
                        filas_para_vista_previa(
                            filas_reporte
                        ),
                        use_container_width=True,
                        hide_index=True,
                    )

                    if st.button(
                        "Actualizar respuestas",
                        use_container_width=True,
                        key="actualizar_reporte_excel",
                    ):
                        st.rerun()

    except requests.exceptions.ConnectionError:
        st.error(
            "No se pudo conectar con FastAPI. "
            "Verifica que Uvicorn esté encendido."
        )

    except requests.exceptions.Timeout:
        st.error(
            "FastAPI tardó demasiado en responder."
        )

    except HttpError as error:
        st.error(
            "Google no permitió consultar las respuestas."
        )
        st.code(str(error))

    except Exception as error:
        st.error(
            "No se pudo generar el reporte."
        )
        st.code(str(error))


elif opcion_menu == "📥 Indicadores":
    st.title("📥 Indicadores")

    st.markdown(
        """
        <p class="descripcion-seccion">
            Resumen de los registros de biopsias almacenados en Google Forms.
        </p>
        """,
        unsafe_allow_html=True,
    )

    try:
        respuesta_formularios = requests.get(
            f"{URL_BACKEND}/formularios",
            timeout=10,
        )
        respuesta_formularios.raise_for_status()

        formularios = respuesta_formularios.json().get(
            "formularios",
            [],
        )

        if not formularios:
            st.info("No hay formularios guardados.")

        else:
            indice_formulario = st.selectbox(
                "Seleccionar formulario",
                options=range(len(formularios)),
                format_func=lambda indice: (
                    f"{formularios[indice].get('titulo', 'Formulario')} "
                    f"— {formularios[indice].get('fecha_creacion', '')}"
                ),
                key="formulario_indicadores_nube",
            )

            formulario_seleccionado = formularios[
                indice_formulario
            ]
            form_id = formulario_seleccionado[
                "form_id_google"
            ]

            with st.spinner(
                "Consultando respuestas de Google Forms..."
            ):
                registros = obtener_registros_formulario(
                    form_id
                )

            if not registros:
                st.info(
                    "Este formulario todavía no tiene respuestas."
                )

            else:
                total_registros = len(registros)

                st.metric("Registros", total_registros)

                # =====================================================
                # CANTIDAD POR PROCEDIMIENTO Y POR BIOPSIA
                # =====================================================

                totales_procedimientos = {
                    opcion: 0
                    for opcion in PROCEDIMIENTOS_BIOPSIA
                }
                totales_biopsias = {
                    opcion: 0
                    for opcion in CATEGORIAS_BIOPSIA
                }
                totales_adicionales = {
                    opcion: 0
                    for opcion in PROCEDIMIENTOS_ADICIONALES
                }

                for registro in registros:
                    # ---------------- PROCEDIMIENTO ----------------
                    if tiene_cantidades_por_opcion_procedimiento(registro):
                        for nombre, cantidad_texto in (
                            obtener_procedimientos_con_cantidad(registro)
                        ):
                            cantidad = convertir_cantidad_biopsia(
                                cantidad_texto
                            )
                            if nombre not in totales_procedimientos:
                                totales_procedimientos[nombre] = 0
                            totales_procedimientos[nombre] += cantidad
                    else:
                        # Formulario nuevo (selección única) o muy antiguo.
                        procedimiento = resolver_procedimiento_biopsia(
                            registro
                        )
                        cantidad_procedimiento = convertir_cantidad_biopsia(
                            registro.get("Cantidad", "")
                        )
                        if procedimiento:
                            if procedimiento not in totales_procedimientos:
                                totales_procedimientos[procedimiento] = 0
                            totales_procedimientos[procedimiento] += (
                                cantidad_procedimiento
                            )

                    # ---------------- ADICIONALES ----------------
                    # (APL, ELVE, enema, inyectoterapia, clip,
                    # polipectomías, mucosectomía). No existe en
                    # formularios antiguos, así que no hay compatibilidad
                    # que resolver aquí.
                    if tiene_cantidades_por_opcion_adicional(registro):
                        for nombre, cantidad_texto in (
                            obtener_adicionales_con_cantidad(registro)
                        ):
                            cantidad = convertir_cantidad_biopsia(
                                cantidad_texto
                            )
                            if nombre not in totales_adicionales:
                                totales_adicionales[nombre] = 0
                            totales_adicionales[nombre] += cantidad

                    # ---------------- BIOPSIA ----------------
                    if tiene_cantidades_por_opcion_biopsia(registro):
                        for nombre, cantidad_texto in (
                            obtener_biopsias_con_cantidad(registro)
                        ):
                            cantidad = convertir_cantidad_biopsia(
                                cantidad_texto
                            )
                            if nombre not in totales_biopsias:
                                totales_biopsias[nombre] = 0
                            totales_biopsias[nombre] += cantidad
                        continue

                    nombre = resolver_nombre_biopsia(registro)
                    cantidad = convertir_cantidad_biopsia(
                        registro.get("Cantidad", "")
                    )

                    if nombre:
                        if nombre not in totales_biopsias:
                            totales_biopsias[nombre] = 0
                        totales_biopsias[nombre] += cantidad

                    # Compatibilidad con formularios anteriores de 3 pares.
                    if not nombre:
                        for numero_biopsia in range(1, 4):
                            nombre_anterior = str(
                                registro.get(
                                    f"Biopsia {numero_biopsia} - Nombre",
                                    "",
                                ) or ""
                            ).strip()
                            cantidad_anterior = convertir_cantidad_biopsia(
                                registro.get(
                                    f"Biopsia {numero_biopsia} - Cantidad",
                                    "",
                                )
                            )
                            if nombre_anterior:
                                if nombre_anterior not in totales_biopsias:
                                    totales_biopsias[nombre_anterior] = 0
                                totales_biopsias[nombre_anterior] += (
                                    cantidad_anterior
                                )

                datos_procedimientos = pd.DataFrame(
                    [
                        {
                            "Procedimiento": nombre,
                            "Cantidad": cantidad,
                        }
                        for nombre, cantidad in totales_procedimientos.items()
                    ]
                )

                st.subheader("Cantidad por procedimiento")
                st.bar_chart(
                    datos_procedimientos,
                    x="Procedimiento",
                    y="Cantidad",
                    use_container_width=True,
                )

                datos_biopsias = pd.DataFrame(
                    [
                        {
                            "Biopsia": nombre,
                            "Cantidad": cantidad,
                        }
                        for nombre, cantidad in totales_biopsias.items()
                    ]
                )

                st.subheader("Cantidad por biopsia")
                st.bar_chart(
                    datos_biopsias,
                    x="Biopsia",
                    y="Cantidad",
                    use_container_width=True,
                )

                datos_adicionales = pd.DataFrame(
                    [
                        {
                            "Procedimiento adicional": nombre,
                            "Cantidad": cantidad,
                        }
                        for nombre, cantidad in totales_adicionales.items()
                    ]
                )

                st.subheader("Cantidad por procedimiento adicional")
                st.bar_chart(
                    datos_adicionales,
                    x="Procedimiento adicional",
                    y="Cantidad",
                    use_container_width=True,
                )

                if st.button(
                    "Actualizar indicadores",
                    use_container_width=True,
                    key="actualizar_indicadores_nube",
                ):
                    st.rerun()

    except requests.exceptions.ConnectionError:
        st.error(
            "No se pudo conectar con FastAPI. "
            "Verifica que Uvicorn esté encendido."
        )

    except requests.exceptions.Timeout:
        st.error(
            "FastAPI tardó demasiado en responder."
        )

    except HttpError as error:
        st.error(
            "Google no permitió consultar las respuestas."
        )
        st.code(str(error))

    except Exception as error:
        st.error(
            "No se pudieron generar los indicadores."
        )
        st.code(str(error))
